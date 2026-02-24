"""
Модуль для экспорта данных в Excel с форматированием
"""
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment


def get_rgb_color_for_excel(val, min_val, max_val, mean_val, is_diagonal=False):
    """Возвращает RGB цвет для значения - четкий градиент от красного к желтому, от желтого к зеленому.
    
    Args:
        val: значение для форматирования
        min_val: минимальное значение
        max_val: максимальное значение
        mean_val: среднее значение
        is_diagonal: флаг диагонального элемента
        
    Returns:
        tuple: (r, g, b) RGB цвет
    """
    # Диагональные значения - белый фон
    if is_diagonal:
        return (255, 255, 255)  # белый
    
    if pd.isna(val) or val == 0:
        return (255, 255, 255)  # белый
    
    # Если значение меньше или равно среднему - градиент от красного к желтому
    if val <= mean_val:
        # Градиент от красного (255,0,0) к желтому (255,255,0)
        if mean_val == min_val:
            ratio = 1.0  # Все значения равны минимуму, делаем желтым
        else:
            ratio = (val - min_val) / (mean_val - min_val)
            ratio = max(0, min(1, ratio))  # Ограничиваем от 0 до 1
        
        # Красный -> Желтый: R=255 постоянный, G растет от 0 до 255, B=0 постоянный
        r = 255
        g = int(255 * ratio)  # от 0 до 255
        b = 0
    else:
        # Градиент от желтого (255,255,0) к зеленому (0,255,0)
        if max_val == mean_val:
            ratio = 1.0  # Все значения равны среднему, делаем зеленым
        else:
            ratio = (val - mean_val) / (max_val - mean_val)
            ratio = max(0, min(1, ratio))  # Ограничиваем от 0 до 1
        
        # Желтый -> Зеленый: R убывает от 255 до 0, G=255 постоянный, B=0 постоянный
        r = int(255 * (1 - ratio))  # от 255 до 0
        g = 255
        b = 0
    
    return (r, g, b)


def apply_excel_color_formatting(worksheet, df, hide_zeros=False):
    """Применяет цветовое форматирование к Excel файлу.
    
    Args:
        worksheet: лист Excel
        df: DataFrame для форматирования
        hide_zeros: если True, нулевые значения скрываются (пустая ячейка)
    """
    min_val = df.min().min()
    max_val = df.max().max()
    mean_val = df.mean().mean()
    
    # Применяем форматирование к данным (начиная со строки 2, т.к. строка 1 - заголовки)
    period_indices_excel = {period: idx for idx, period in enumerate(df.index)}
    
    # Определяем, на какой строке начинаются данные (обычно строка 2, если есть заголовок индекса)
    start_row = 2  # Начальная строка с данными (строка 1 - заголовки столбцов и индекса)
    
    for row_idx, period in enumerate(df.index, start=start_row):
        for col_idx, col_period in enumerate(df.columns, start=2):  # Столбец 1 - индекс, данные с столбца 2
            cell = worksheet.cell(row=row_idx, column=col_idx)
            value = df.loc[period, col_period]
            
            # Проверяем, является ли это диагональю
            is_diagonal = (period == col_period)
            
            if is_diagonal:
                # Диагональ - белый фон, жирный шрифт
                r, g, b = get_rgb_color_for_excel(value, min_val, max_val, mean_val, is_diagonal=True)
                hex_color = f"{r:02X}{g:02X}{b:02X}"
                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                cell.font = Font(color="000000", bold=True)  # чёрный текст, жирный
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif not pd.isna(value) and value != 0:
                r, g, b = get_rgb_color_for_excel(value, min_val, max_val, mean_val, is_diagonal=False)
                # Формат RGB для openpyxl: RRGGBB
                hex_color = f"{r:02X}{g:02X}{b:02X}"
                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                cell.font = Font(color="000000")  # чёрный текст
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                # Нулевые значения или пустые
                if hide_zeros and not is_diagonal:
                    # Скрываем нули (пустая ячейка)
                    cell.value = ""
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    cell.font = Font(color="FFFFFF")  # белый текст на белом фоне
                else:
                    # Показываем нули
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    cell.font = Font(color="000000")
                cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_excel_cohort_formatting(worksheet, df, sorted_periods):
    """Применяет цветовое форматирование с горизонтальной динамикой к Excel файлу для таблицы когорт.
    
    Args:
        worksheet: лист Excel
        df: DataFrame для форматирования
        sorted_periods: отсортированный список периодов
    """
    period_indices = {period: idx for idx, period in enumerate(sorted_periods)}
    
    # Для горизонтальной динамики рассчитываем min/max/mean для каждой строки отдельно
    def get_row_stats(row_period):
        row_idx = period_indices.get(row_period, 0)
        row_values = []
        for col_period in df.columns:
            col_idx = period_indices.get(col_period, 0)
            # Учитываем только значения после диагонали
            if row_period != col_period and col_idx >= row_idx:
                val = df.loc[row_period, col_period]
                if not pd.isna(val) and val > 0:
                    row_values.append(val)
        if row_values:
            return min(row_values), max(row_values), sum(row_values) / len(row_values)
        return 0, 0, 0
    
    start_row = 2
    for row_idx, period in enumerate(df.index, start=start_row):
        row_period_idx = period_indices.get(period, 0)
        row_min, row_max, row_mean = get_row_stats(period)
        
        for col_idx, col_period in enumerate(df.columns, start=2):
            col_period_idx = period_indices.get(col_period, 0)
            cell = worksheet.cell(row=row_idx, column=col_idx)
            value = df.loc[period, col_period]
            is_diagonal = (period == col_period)
            
            # Скрываем значения до диагонали
            if not is_diagonal and col_period_idx < row_period_idx:
                cell.value = ""
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                cell.font = Font(color="FFFFFF")  # белый текст на белом фоне
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif is_diagonal:
                # Диагональ - белый фон, жирный шрифт
                r, g, b = get_rgb_color_for_excel(value, row_min, row_max, row_mean, is_diagonal=True)
                hex_color = f"{r:02X}{g:02X}{b:02X}"
                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                cell.font = Font(color="000000", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # Форматируем как целое число
                if cell.value is not None and not isinstance(cell.value, str):
                    cell.number_format = '0'
            elif not pd.isna(value) and value > 0:
                r, g, b = get_rgb_color_for_excel(value, row_min, row_max, row_mean, is_diagonal=False)
                hex_color = f"{r:02X}{g:02X}{b:02X}"
                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                cell.font = Font(color="000000")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # Форматируем как целое число
                if cell.value is not None and not isinstance(cell.value, str):
                    cell.number_format = '0'
            else:
                # Нулевые значения и пустые: скрываем (пустая ячейка, белый фон)
                cell.value = ""
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                cell.font = Font(color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_excel_percent_formatting(worksheet, df, sorted_periods):
    """Применяет цветовое форматирование и форматирование процентов к Excel файлу для таблицы накопления в %.
    
    Args:
        worksheet: лист Excel
        df: DataFrame для форматирования
        sorted_periods: отсортированный список периодов
    """
    period_indices = {period: idx for idx, period in enumerate(sorted_periods)}
    
    # Для горизонтальной динамики рассчитываем min/max/mean для каждой строки отдельно
    def get_row_stats(row_period):
        row_idx = period_indices.get(row_period, 0)
        row_values = []
        for col_period in df.columns:
            col_idx = period_indices.get(col_period, 0)
            if row_period != col_period and col_idx >= row_idx:
                val = df.loc[row_period, col_period]
                if not pd.isna(val) and val > 0:
                    row_values.append(val)
        if row_values:
            return min(row_values), max(row_values), sum(row_values) / len(row_values)
        return 0, 0, 0
    
    start_row = 2
    for row_idx, period in enumerate(df.index, start=start_row):
        row_period_idx = period_indices.get(period, 0)
        row_min, row_max, row_mean = get_row_stats(period)
        
        for col_idx, col_period in enumerate(df.columns, start=2):
            col_period_idx = period_indices.get(col_period, 0)
            cell = worksheet.cell(row=row_idx, column=col_idx)
            value = df.loc[period, col_period]
            is_diagonal = (period == col_period)
            
            # Скрываем значения до диагонали
            if not is_diagonal and col_period_idx < row_period_idx:
                cell.value = ""
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                cell.font = Font(color="FFFFFF")  # белый текст на белом фоне
            elif is_diagonal:
                # Диагональ - 100.0% (сохраняем как число 1.0, Excel покажет как 100%)
                cell.value = 1.0
                cell.number_format = '0.0%'  # Процентный формат Excel
                r, g, b = get_rgb_color_for_excel(100.0, row_min, row_max, row_mean, is_diagonal=True)
                hex_color = f"{r:02X}{g:02X}{b:02X}"
                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                cell.font = Font(color="000000", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif not pd.isna(value) and value > 0:
                # Сохраняем как число (value уже в процентах, конвертируем в долю для Excel)
                cell.value = value / 100.0  # Конвертируем проценты в долю (45.7 -> 0.457)
                cell.number_format = '0.0%'  # Процентный формат Excel
                r, g, b = get_rgb_color_for_excel(value, row_min, row_max, row_mean, is_diagonal=False)
                hex_color = f"{r:02X}{g:02X}{b:02X}"
                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                cell.font = Font(color="000000")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.value = ""
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                cell.font = Font(color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_excel_inflow_formatting(worksheet, df, sorted_periods):
    """Применяет цветовое форматирование и форматирование процентов к Excel файлу для таблицы притока в %.
    
    Args:
        worksheet: лист Excel
        df: DataFrame для форматирования
        sorted_periods: отсортированный список периодов
    """
    period_indices = {period: idx for idx, period in enumerate(sorted_periods)}
    
    # Для горизонтальной динамики рассчитываем min/max/mean для каждой строки отдельно
    def get_row_stats(row_period):
        row_idx = period_indices.get(row_period, 0)
        row_values = []
        for col_period in df.columns:
            col_idx = period_indices.get(col_period, 0)
            if row_period != col_period and col_idx >= row_idx:
                val = df.loc[row_period, col_period]
                if not pd.isna(val) and val > 0:
                    row_values.append(val)
        if row_values:
            return min(row_values), max(row_values), sum(row_values) / len(row_values)
        return 0, 0, 0
    
    start_row = 2
    for row_idx, period in enumerate(df.index, start=start_row):
        row_period_idx = period_indices.get(period, 0)
        row_min, row_max, row_mean = get_row_stats(period)
        
        for col_idx, col_period in enumerate(df.columns, start=2):
            col_period_idx = period_indices.get(col_period, 0)
            cell = worksheet.cell(row=row_idx, column=col_idx)
            value = df.loc[period, col_period]
            is_diagonal = (period == col_period)
            
            # Скрываем значения до диагонали
            if not is_diagonal and col_period_idx < row_period_idx:
                cell.value = ""
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                cell.font = Font(color="FFFFFF")  # белый текст на белом фоне
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif is_diagonal:
                # Диагональ - 0.0% (сохраняем как число 0.0, Excel покажет как 0.0%)
                cell.value = 0.0
                cell.number_format = '0.0%'  # Процентный формат Excel
                r, g, b = get_rgb_color_for_excel(0.0, row_min, row_max, row_mean, is_diagonal=True)
                hex_color = f"{r:02X}{g:02X}{b:02X}"
                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                cell.font = Font(color="000000", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif not pd.isna(value) and value > 0:
                # Сохраняем как число (value уже в процентах, конвертируем в долю для Excel)
                cell.value = value / 100.0  # Конвертируем проценты в долю (45.7 -> 0.457)
                cell.number_format = '0.0%'  # Процентный формат Excel
                r, g, b = get_rgb_color_for_excel(value, row_min, row_max, row_mean, is_diagonal=False)
                hex_color = f"{r:02X}{g:02X}{b:02X}"
                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                cell.font = Font(color="000000")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.value = ""
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                cell.font = Font(color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")

