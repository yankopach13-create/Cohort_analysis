import re
import streamlit as st
import pandas as pd
import numpy as np
import io
import os
import platform
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Используем неинтерактивный бэкенд
import seaborn as sns
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Импорты из новых модулей
from config import PAGE_CONFIG, TEMPLATE_IMAGE_PATHS, CATEGORIES_TEMPLATE_IMAGE_PATHS
from utils import parse_period, parse_year_month, create_copy_button, detect_columns
try:
    from utils import get_period_after_label
except ImportError:
    def get_period_after_label(sorted_periods):
        """Запасной вариант, если в utils нет функции (старая версия на Cloud)."""
        return 'месяца'
from data_processing import (
    get_cohort_clients, get_accumulation_clients, get_client_cohorts,
    get_churn_clients, get_inflow_clients, build_churn_table,
    create_period_clients_cache
)
from matrix_builder import (
    build_cohort_matrix, build_accumulation_matrix,
    build_accumulation_percent_matrix, build_inflow_matrix
)
from ui_components import color_gradient, apply_matrix_color_gradient
import inspect
from excel_exporter import (
    apply_excel_color_formatting, apply_excel_cohort_formatting,
    apply_excel_percent_formatting, apply_excel_inflow_formatting
)

def _excel_format_kwargs(fn, data_start_row):
    """Ключевые аргументы для функций форматирования Excel (data_start_row поддерживается не во всех версиях)."""
    kwargs = {}
    if "data_start_row" in inspect.signature(fn).parameters:
        kwargs["data_start_row"] = data_start_row
    return kwargs


def _churn_int(val, default=0):
    """Число из ячейки таблицы оттока (значение '-' для последней когорты → default)."""
    if val == '-' or pd.isna(val):
        return default
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return default


def _churn_float(val, default=0.0):
    """Процент/float из ячейки таблицы оттока (значение '-' → default)."""
    if val == '-' or pd.isna(val):
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


# Настройка страницы
st.set_page_config(**PAGE_CONFIG)

# Функция create_copy_button теперь импортируется из utils

st.title("Когортный анализ, возвращаемость и отток")
st.markdown("---")

# Глобальные CSS стили для всех таблиц (выравнивание по центру)
st.markdown("""
<style>
div[data-testid="stDataFrame"] table,
div[data-testid="stDataFrame"] table th,
div[data-testid="stDataFrame"] table td {
    text-align: center !important;
}
div[data-testid="stDataFrame"] th,
div[data-testid="stDataFrame"] td {
    text-align: center !important;
}
</style>
""", unsafe_allow_html=True)

# Инициализация session state для хранения загруженных данных
if 'uploaded_data' not in st.session_state:
    st.session_state.uploaded_data = None

# Инициализируем флаг загрузки второго файла
if 'categories_file_uploaded' not in st.session_state:
    st.session_state.categories_file_uploaded = False
if 'df' not in st.session_state:
    st.session_state.df = None
if 'cohort_matrix' not in st.session_state:
    st.session_state.cohort_matrix = None
if 'cohort_info' not in st.session_state:
    st.session_state.cohort_info = None
if 'sorted_periods' not in st.session_state:
    st.session_state.sorted_periods = None
if 'year_month_col' not in st.session_state:
    st.session_state.year_month_col = None
if 'client_col' not in st.session_state:
    st.session_state.client_col = None

# Функции parse_period, parse_year_month, color_gradient, apply_matrix_color_gradient теперь импортируются из модулей
# Функции apply_excel_* также импортируются из excel_exporter
# Все функции построения матриц импортируются из matrix_builder
# Все функции обработки данных импортируются из data_processing

# Дублирующиеся функции удалены - они теперь в модулях:
# - apply_excel_color_formatting, apply_excel_cohort_formatting, apply_excel_percent_formatting, apply_excel_inflow_formatting -> excel_exporter.py
# - build_cohort_matrix, build_accumulation_matrix, build_accumulation_percent_matrix, build_inflow_matrix -> matrix_builder.py
# - get_cohort_clients, get_accumulation_clients, get_client_cohorts, get_churn_clients, get_inflow_clients, build_churn_table -> data_processing.py
# Функция загрузки Excel файла
# Создаем колонки для выравнивания заголовков на одном уровне
col_header_left, col_header_right = st.columns([1, 1])

with col_header_left:
    st.header("📊 Когортный анализ")

with col_header_right:
    st.subheader("📋 Шаблон загрузки данных из Qlik")

# Блок шаблона Qlik - инструкции слева, изображение и загрузчик справа
col_template_instructions, col_template_image = st.columns([1, 1])

with col_template_image:
    # Пытаемся найти скриншот шаблона Qlik (изображение вверху)
    qlik_image_paths = [
        'Qlik.png',
        'Qlik.jpg',
        'Qlik.jpeg',
        'qlik_template.png',
        'qlik_template.jpg',
        'qlik_template.jpeg',
        'шаблон_qlik.png',
        'шаблон_qlik.jpg',
        'шаблон_qlik.jpeg',
        'qlik.png',
        'qlik.jpg',
        'qlik.jpeg'
    ]
    image_found = False
    for img_path in qlik_image_paths:
        if os.path.exists(img_path):
            st.image(img_path, use_container_width=True)
            image_found = True
            break
    if not image_found:
        st.info("📸 Поместите скриншот шаблона загрузки данных из Qlik в папку проекта с одним из имён: Qlik.png, qlik_template.png, шаблон_qlik.png или qlik.png")
    
    # Загрузчик Excel файла прямо под изображением
    uploaded_file = st.file_uploader(
        "Выберите Excel файл для загрузки",
        type=['xlsx', 'xls'],
        help="Поддерживаются файлы формата .xlsx и .xls"
    )

with col_template_instructions:
    # Текст инструкций
    st.markdown("""
    1. Зайдите в Qlik, анализ чеков.
    
    2. Отберите анализируемые категории в одном из разрезов Группа1/2/3/4.
    
    3. Отберите анализируемый период.
    
    4. Зайдите на лист "Конструктор" и выведите отчёт по шаблону справа.
    
    Настройте фильтрами построение динамики когорт: Год-Месяц или Год-Неделя.
    
    5. Скачайте документ в Qlik и загрузите в ячейку справа.
    """)

if uploaded_file is not None:
    try:
        # Загрузка Excel файла
        if uploaded_file.name.endswith('.xlsx'):
            df = pd.read_excel(uploaded_file, engine='openpyxl')
        else:
            df = pd.read_excel(uploaded_file, engine='xlrd')
        
        # Сохранение данных в session state
        # Проверяем, новый ли это файл
        is_new_file = (
            st.session_state.uploaded_data is None or 
            st.session_state.uploaded_data.name != uploaded_file.name
        )
        
        st.session_state.uploaded_data = uploaded_file
        st.session_state.df = df
        
        # Очищаем старую информацию только при загрузке нового файла
        if is_new_file:
            st.session_state.cohort_info = None
            st.session_state.cohort_matrix = None
            st.session_state.sorted_periods = None
            st.session_state.year_month_col = None
            st.session_state.client_col = None
        
        # Построение когортной матрицы
        # Уменьшаем отступ перед блоком матриц
        st.markdown("<div style='margin-top: 10px;'></div>", unsafe_allow_html=True)
        
        # Определяем столбцы автоматически
        year_month_col, client_col = detect_columns(df)
        
        # Если столбцы не найдены, показываем ошибку
        if year_month_col is None:
            st.error("❌ Не найден столбец с периодом (Год-месяц или Год-Неделя). Убедитесь, что в файле есть столбец с названием, содержащим 'Год' и 'месяц' или 'неделя'.")
            st.stop()
        
        if client_col is None:
            st.error("❌ Не найден столбец с кодом клиента. Убедитесь, что в файле есть столбец с названием, содержащим 'Код' и 'клиент'.")
            st.stop()
        
        # Сохраняем выбранные столбцы в session state
        st.session_state.year_month_col = year_month_col
        st.session_state.client_col = client_col
        
        # Построение матрицы
        if year_month_col and client_col:
            try:
                # Проверяем, есть ли уже вычисленные данные
                need_recompute = (
                    st.session_state.cohort_matrix is None or
                    st.session_state.sorted_periods is None or
                    st.session_state.year_month_col != year_month_col or
                    st.session_state.client_col != client_col
                )
                
                # Создаём контейнер для всего контента
                content_placeholder = st.empty()
                
                if need_recompute:
                    # Единый спиннер для всех расчётов - показываем только его
                    with content_placeholder.container():
                        with st.spinner("Расчёт и анализ данных..."):
                            # Построение когортной матрицы
                            cohort_matrix, sorted_periods = build_cohort_matrix(
                                df, 
                                year_month_col, 
                                client_col, 
                                value_type='clients'
                            )
                            st.session_state.cohort_matrix = cohort_matrix
                            st.session_state.sorted_periods = sorted_periods
                            st.session_state.period_after_label = get_period_after_label(sorted_periods)

                            # Кэшируем множества клиентов по периодам для быстрого доступа в функциях получения клиентов
                            # Оптимизация: используем groupby вместо циклов с фильтрацией
                            period_clients_cache = {}
                            df_filtered = df[[year_month_col, client_col]].dropna()
                            for period, group in df_filtered.groupby(year_month_col):
                                period_clients_cache[period] = set(group[client_col].unique())
                            # Добавляем пустые множества для периодов без клиентов
                            for period in sorted_periods:
                                if period not in period_clients_cache:
                                    period_clients_cache[period] = set()
                            st.session_state.period_clients_cache = period_clients_cache
                            
                            # Кэшируем когорты клиентов (первый период появления каждого клиента)
                            client_cohorts_cache = get_client_cohorts(df, year_month_col, client_col, sorted_periods)
                            st.session_state.client_cohorts_cache = client_cohorts_cache
                            
                            # Вычисляем статистику по диагонали (количество клиентов в каждом периоде)
                            diagonal_values = {period: cohort_matrix.loc[period, period] for period in sorted_periods}
                            
                            # Находим максимум и минимум
                            max_clients = max(diagonal_values.values())
                            min_clients = min(diagonal_values.values())
                            max_period = [period for period, val in diagonal_values.items() if val == max_clients][0]
                            min_period = [period for period, val in diagonal_values.items() if val == min_clients][0]
                            
                            # Первый и последний период
                            first_period = sorted_periods[0]
                            last_period = sorted_periods[-1]
                            
                            # Сохраняем информацию в session state для отображения в правой колонке
                            st.session_state.cohort_info = {
                                'num_periods': len(sorted_periods),
                                'first_period': first_period,
                                'last_period': last_period,
                                'max_clients': max_clients,
                                'max_period': max_period,
                                'min_clients': min_clients,
                                'min_period': min_period
                            }
                            
                            # Построение всех остальных матриц внутри спиннера
                            st.session_state.accumulation_matrix = build_accumulation_matrix(df, year_month_col, client_col, sorted_periods)
                            st.session_state.accumulation_percent_matrix = build_accumulation_percent_matrix(st.session_state.accumulation_matrix, cohort_matrix)
                            st.session_state.inflow_matrix = build_inflow_matrix(st.session_state.accumulation_percent_matrix)
                            
                            # Кэшируем когорты клиентов (первый период появления каждого клиента)
                            client_cohorts_cache = get_client_cohorts(df, year_month_col, client_col, sorted_periods)
                            st.session_state.client_cohorts_cache = client_cohorts_cache
                            
                            st.session_state.churn_table = build_churn_table(df, year_month_col, client_col, sorted_periods, cohort_matrix, st.session_state.accumulation_matrix, st.session_state.accumulation_percent_matrix, client_cohorts_cache, period_clients_cache)
                            
                            # Кэшируем множества клиентов по периодам для быстрого доступа в функциях получения клиентов
                            # Оптимизация: используем groupby вместо циклов с фильтрацией
                            period_clients_cache = {}
                            df_filtered = df[[year_month_col, client_col]].dropna()
                            for period, group in df_filtered.groupby(year_month_col):
                                period_clients_cache[period] = set(group[client_col].unique())
                            # Добавляем пустые множества для периодов без клиентов
                            for period in sorted_periods:
                                if period not in period_clients_cache:
                                    period_clients_cache[period] = set()
                            st.session_state.period_clients_cache = period_clients_cache
                    
                    # После завершения всех расчётов очищаем placeholder и отображаем весь контент
                    content_placeholder.empty()
                else:
                    # Используем сохраненные данные
                    cohort_matrix = st.session_state.cohort_matrix
                    sorted_periods = st.session_state.sorted_periods
                    # Проверяем наличие остальных матриц
                    if st.session_state.get('accumulation_matrix') is None:
                        st.session_state.accumulation_matrix = build_accumulation_matrix(df, year_month_col, client_col, sorted_periods)
                    if st.session_state.get('accumulation_percent_matrix') is None:
                        st.session_state.accumulation_percent_matrix = build_accumulation_percent_matrix(st.session_state.accumulation_matrix, cohort_matrix)
                    if st.session_state.get('inflow_matrix') is None:
                        st.session_state.inflow_matrix = build_inflow_matrix(st.session_state.accumulation_percent_matrix)
                    if st.session_state.get('churn_table') is None:
                        client_cohorts_cache = st.session_state.get('client_cohorts_cache')
                        if client_cohorts_cache is None:
                            client_cohorts_cache = get_client_cohorts(df, year_month_col, client_col, sorted_periods)
                            st.session_state.client_cohorts_cache = client_cohorts_cache
                        period_clients_cache = st.session_state.get('period_clients_cache')
                        st.session_state.churn_table = build_churn_table(df, year_month_col, client_col, sorted_periods, cohort_matrix, st.session_state.accumulation_matrix, st.session_state.accumulation_percent_matrix, client_cohorts_cache, period_clients_cache)
                    
                    # Создаем кэш множеств клиентов, если его еще нет
                    if st.session_state.get('period_clients_cache') is None:
                        # Оптимизация: используем groupby вместо циклов с фильтрацией
                        period_clients_cache = {}
                        df_filtered = df[[year_month_col, client_col]].dropna()
                        for period, group in df_filtered.groupby(year_month_col):
                            period_clients_cache[period] = set(group[client_col].unique())
                        # Добавляем пустые множества для периодов без клиентов
                        for period in sorted_periods:
                            if period not in period_clients_cache:
                                period_clients_cache[period] = set()
                        st.session_state.period_clients_cache = period_clients_cache
                    
                    # Создаем кэш когорт клиентов, если его еще нет
                    if st.session_state.get('client_cohorts_cache') is None:
                        client_cohorts_cache = get_client_cohorts(df, year_month_col, client_col, sorted_periods)
                        st.session_state.client_cohorts_cache = client_cohorts_cache
                    if st.session_state.get('period_after_label') is None:
                        st.session_state.period_after_label = get_period_after_label(sorted_periods)
                
                # Получаем информацию из session state
                info = st.session_state.cohort_info
                
                # Отображаем кнопки скачивания под блоком загрузки (горизонтально)
                if info:
                        # Создаем функцию для генерации полного отчёта
                        def create_full_report_excel():
                            """Создает полный Excel отчёт со всеми таблицами"""
                            buffer = io.BytesIO()
                            
                            # Импорты для форматирования Excel
                            from openpyxl.styles import Font
                            from openpyxl.utils import get_column_letter
                            
                            # Получаем данные из session state
                            cohort_matrix = st.session_state.cohort_matrix
                            sorted_periods = st.session_state.sorted_periods
                            df = st.session_state.df
                            year_month_col = st.session_state.year_month_col
                            client_col = st.session_state.client_col
                            # Подпись «Продукт построения когорт» из первого столбца первого документа
                            product_col = df.columns[0] if df is not None and len(df.columns) > 0 else None
                            if product_col is not None:
                                _up = sorted(df[product_col].dropna().astype(str).str.strip().unique())
                                _up = [p for p in _up if p]
                                products_label = ", ".join(_up) if _up else ""
                            else:
                                products_label = ""
                            # Смещение строки данных при наличии заголовка «Продукт построения когорт»
                            data_start_row = 4 if products_label else 2
                            table_startrow = 2 if products_label else 0
                            # Если второй файл загружен, но данные ещё не обработаны, обрабатываем их на лету
                            uploaded_file_categories = st.session_state.get('upload_categories_file')
                            if uploaded_file_categories is not None and ('df_categories' not in st.session_state or st.session_state.df_categories is None):
                                try:
                                    # Загружаем и обрабатываем файл на лету
                                    if uploaded_file_categories.name.endswith('.xlsx'):
                                        df_categories_temp = pd.read_excel(uploaded_file_categories, engine='openpyxl')
                                    else:
                                        df_categories_temp = pd.read_excel(uploaded_file_categories, engine='xlrd')
                                    
                                    # Определяем столбцы (упрощённая версия)
                                    group_col_temp = None
                                    year_month_col_temp = None
                                    client_code_col_temp = None
                                    
                                    for col in df_categories_temp.columns:
                                        col_lower = str(col).lower().strip()
                                        if 'группа' in col_lower and group_col_temp is None:
                                            group_col_temp = col
                                        if (('год' in col_lower and ('месяц' in col_lower or 'неделя' in col_lower)) or ('год-месяц' in col_lower) or ('год-неделя' in col_lower)) and year_month_col_temp is None:
                                            year_month_col_temp = col
                                        if 'код' in col_lower and 'клиент' in col_lower and client_code_col_temp is None:
                                            client_code_col_temp = col
                                    
                                    if group_col_temp and client_code_col_temp:
                                        categories_temp = sorted([str(cat) for cat in df_categories_temp[group_col_temp].dropna().unique() if str(cat).strip() != ''])
                                        
                                        # Сохраняем базовые данные
                                        st.session_state.df_categories = df_categories_temp
                                        st.session_state.categories_list = categories_temp
                                        st.session_state.group_col_name = group_col_temp
                                        st.session_state.year_month_col_name = year_month_col_temp
                                        st.session_state.client_code_col_name = client_code_col_temp
                                        
                                        # Полная обработка данных для создания category_summary_table
                                        if 'churn_table' in st.session_state and st.session_state.churn_table is not None:
                                            # Получаем необходимые данные
                                            period_clients_cache = st.session_state.get('period_clients_cache', None)
                                            client_cohorts_cache = st.session_state.get('client_cohorts_cache', None)
                                            churn_table = st.session_state.churn_table
                                            
                                            # Рассчитываем метрики для всех когорт
                                            total_present_after_cohort_by_cohort = {}
                                            total_present_after_cohort_percent_by_cohort = {}
                                            network_churn_by_cohort = {}
                                            network_churn_percent_by_cohort = {}
                                            
                                            for cohort_period in sorted_periods:
                                                # Получаем клиентов оттока для этой когорты
                                                churn_clients_set_cohort = set(get_churn_clients(df, year_month_col, client_col, sorted_periods, cohort_period, period_clients_cache, client_cohorts_cache))
                                                churn_clients_set_cohort = {str(client) for client in churn_clients_set_cohort}
                                                
                                                # Получаем отток из категории для этой когорты
                                                cohort_row = churn_table[churn_table['Когорта'] == cohort_period]
                                                churn_count_cohort = _churn_int(cohort_row.iloc[0]['Отток кол-во']) if not cohort_row.empty else 0
                                                cohort_size_cohort = int(cohort_row.iloc[0]['Кол-во клиентов когорты']) if not cohort_row.empty else 0
                                                
                                                # Определяем периоды ПОСЛЕ когорты
                                                cohort_index_cohort = sorted_periods.index(cohort_period) if cohort_period in sorted_periods else 0
                                                periods_from_cohort_cohort = sorted_periods[cohort_index_cohort:]
                                                periods_after_cohort_cohort = periods_from_cohort_cohort[1:] if len(periods_from_cohort_cohort) > 1 else []
                                                
                                                # Клиенты оттока, присутствующие в других категориях ПОСЛЕ месяца когорты
                                                all_category_clients_after_cohort = set()
                                                if year_month_col_temp is not None and len(periods_after_cohort_cohort) > 0:
                                                    for category in categories_temp:
                                                        category_data = df_categories_temp[df_categories_temp[group_col_temp] == category]
                                                        category_data_filtered = category_data[category_data[year_month_col_temp].isin(periods_after_cohort_cohort)]
                                                        category_clients = set(category_data_filtered[client_code_col_temp].dropna().astype(str).unique())
                                                        all_category_clients_after_cohort.update(category_clients)
                                                elif year_month_col_temp is None:
                                                    # Если нет столбца год-месяц, используем всех клиентов из категорий
                                                    for category in categories_temp:
                                                        category_data = df_categories_temp[df_categories_temp[group_col_temp] == category]
                                                        category_clients = set(category_data[client_code_col_temp].dropna().astype(str).unique())
                                                        all_category_clients_after_cohort.update(category_clients)
                                                
                                                present_in_categories_after_cohort = churn_clients_set_cohort & all_category_clients_after_cohort
                                                total_present_after_cohort_by_cohort[cohort_period] = len(present_in_categories_after_cohort)
                                                
                                                # % присутствия после месяца когорты
                                                present_after_cohort_percent = (len(present_in_categories_after_cohort) / cohort_size_cohort * 100) if cohort_size_cohort > 0 else 0
                                                total_present_after_cohort_percent_by_cohort[cohort_period] = present_after_cohort_percent
                                                
                                                # Отток из сети
                                                network_churn_cohort = churn_count_cohort - len(present_in_categories_after_cohort)
                                                network_churn_by_cohort[cohort_period] = max(0, network_churn_cohort)
                                                
                                                # % оттока из сети
                                                network_churn_percent_cohort = (network_churn_by_cohort[cohort_period] / cohort_size_cohort * 100) if cohort_size_cohort > 0 else 0
                                                network_churn_percent_by_cohort[cohort_period] = network_churn_percent_cohort
                                            
                                            # Ключи метрик с учётом типа периода (недели/месяцы)
                                            _pa = st.session_state.get('period_after_label', 'месяца')
                                            _key_итого = f"Итого присутствуют в других категориях после {_pa} когорты"
                                            _key_доля = f"Доля присутствуют в других категориях после {_pa} когорты"
                                            summary_table_excel = pd.DataFrame({
                                                'Отток из сети': network_churn_by_cohort,
                                                'Доля оттока из сети от когорты': network_churn_percent_by_cohort,
                                                _key_итого: total_present_after_cohort_by_cohort,
                                                _key_доля: total_present_after_cohort_percent_by_cohort
                                            })
                                            summary_table_excel = summary_table_excel.T
                                            
                                            # Сохраняем category_summary_table
                                            st.session_state.category_summary_table = summary_table_excel
                                            st.session_state.category_cohort_table = None
                                except Exception as e:
                                    # Если не удалось обработать на лету, просто пропускаем таблицу 6
                                    pass
                        
                            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                                workbook = writer.book
                                
                                # Получаем все матрицы
                                accumulation_matrix = build_accumulation_matrix(df, year_month_col, client_col, sorted_periods)
                                accumulation_percent_matrix = build_accumulation_percent_matrix(accumulation_matrix, cohort_matrix)
                                inflow_matrix = build_inflow_matrix(accumulation_percent_matrix)
                                
                                # Таблица 1: Динамика уникальных клиентов когорт
                                cohort_matrix_copy = cohort_matrix.copy()
                                cohort_matrix_copy.index.name = 'Когорта / Период'
                                cohort_matrix_copy.to_excel(writer, sheet_name="1. Динамика уникальных клиентов", startrow=table_startrow, index=True)
                                worksheet1 = writer.sheets["1. Динамика уникальных клиентов"]
                                if products_label:
                                    worksheet1.cell(row=1, column=1, value=f"Продукт построения когорт: {products_label}")
                                    worksheet1.merge_cells(f"A1:{get_column_letter(1 + len(cohort_matrix.columns))}1")
                                    worksheet1.cell(row=1, column=1).font = Font(bold=True, size=11)
                                apply_excel_cohort_formatting(worksheet1, cohort_matrix.astype(float), sorted_periods, **_excel_format_kwargs(apply_excel_cohort_formatting, data_start_row))
                                
                                # Таблица 2: Динамика накопления возврата
                                accumulation_matrix_copy = accumulation_matrix.copy()
                                accumulation_matrix_copy.index.name = 'Когорта / Период'
                                accumulation_matrix_copy.to_excel(writer, sheet_name="2. Динамика накопления", startrow=table_startrow, index=True)
                                worksheet2 = writer.sheets["2. Динамика накопления"]
                                if products_label:
                                    worksheet2.cell(row=1, column=1, value=f"Продукт построения когорт: {products_label}")
                                    worksheet2.merge_cells(f"A1:{get_column_letter(1 + len(accumulation_matrix.columns))}1")
                                    worksheet2.cell(row=1, column=1).font = Font(bold=True, size=11)
                                apply_excel_color_formatting(worksheet2, accumulation_matrix.astype(float), hide_zeros=True, **_excel_format_kwargs(apply_excel_color_formatting, data_start_row))
                                # Форматируем значения как целые числа (только для непустых ячеек)
                                for row_idx in range(data_start_row, data_start_row + len(accumulation_matrix.index)):
                                    for col_idx in range(2, len(accumulation_matrix.columns) + 2):
                                        cell = worksheet2.cell(row=row_idx, column=col_idx)
                                        if cell.value is not None and not isinstance(cell.value, str) and cell.value != "":
                                            cell.number_format = '0'  # Формат целого числа
                                
                                # Таблица 3: Динамика накопления возврата в %
                                accumulation_percent_matrix_copy = accumulation_percent_matrix.copy()
                                accumulation_percent_matrix_copy.index.name = 'Когорта / Период'
                                accumulation_percent_matrix_copy.to_excel(writer, sheet_name="3. Динамика накопления %", startrow=table_startrow, index=True)
                                worksheet3 = writer.sheets["3. Динамика накопления %"]
                                if products_label:
                                    worksheet3.cell(row=1, column=1, value=f"Продукт построения когорт: {products_label}")
                                    worksheet3.merge_cells(f"A1:{get_column_letter(1 + len(accumulation_percent_matrix.columns))}1")
                                    worksheet3.cell(row=1, column=1).font = Font(bold=True, size=11)
                                apply_excel_percent_formatting(worksheet3, accumulation_percent_matrix, sorted_periods, **_excel_format_kwargs(apply_excel_percent_formatting, data_start_row))
                                
                                # Таблица 4: Приток возврата в %
                                inflow_matrix_copy = inflow_matrix.copy()
                                inflow_matrix_copy.index.name = 'Когорта / Период'
                                inflow_matrix_copy.to_excel(writer, sheet_name="4. Приток возврата %", startrow=table_startrow, index=True)
                                worksheet4 = writer.sheets["4. Приток возврата %"]
                                if products_label:
                                    worksheet4.cell(row=1, column=1, value=f"Продукт построения когорт: {products_label}")
                                    worksheet4.merge_cells(f"A1:{get_column_letter(1 + len(inflow_matrix.columns))}1")
                                    worksheet4.cell(row=1, column=1).font = Font(bold=True, size=11)
                                apply_excel_inflow_formatting(worksheet4, inflow_matrix, sorted_periods, **_excel_format_kwargs(apply_excel_inflow_formatting, data_start_row))
                                
                                # Таблица 5: Отток клиентов из категории
                                churn_table_full = build_churn_table(df, year_month_col, client_col, sorted_periods, cohort_matrix, accumulation_matrix, accumulation_percent_matrix, None, None)
                                churn_table_copy = churn_table_full.copy()
                                # Не конвертируем проценты в строки - сохраняем как числа для возможности расчетов
                                churn_table_copy.to_excel(writer, sheet_name="5. Отток клиентов из категории", startrow=0, index=False)
                                worksheet5 = writer.sheets["5. Отток клиентов из категории"]
                                # Форматируем значения: числа как целые, проценты как проценты
                                from openpyxl.styles import Alignment as ExcelAlignment
                                for row_idx in range(2, len(churn_table_copy) + 2):
                                    for col_idx in range(1, len(churn_table_copy.columns) + 1):
                                        cell = worksheet5.cell(row=row_idx, column=col_idx)
                                        cell.alignment = ExcelAlignment(horizontal="center", vertical="center")
                                        col_name = churn_table_copy.columns[col_idx - 1]
                                        if col_name in ['Кол-во клиентов когорты', 'Накопительное кол-во возврата', 'Отток кол-во']:
                                            # Колонки с числами
                                            if cell.value is not None and not isinstance(cell.value, str):
                                                cell.number_format = '0'  # Формат целого числа
                                        elif col_name in ['Накопительный % возврата', 'Отток %']:
                                            # Колонки с процентами - сохраняем как число (уже в процентах, конвертируем в долю)
                                            if cell.value is not None and not isinstance(cell.value, str):
                                                # Значение уже в процентах (например, 45.7), конвертируем в долю (0.457)
                                                cell.value = float(cell.value) / 100.0
                                                cell.number_format = '0.0%'  # Процентный формат Excel
                                
                                # Таблица 7: Присутствие клиентов оттока когорты в других категориях товаров
                                # Проверяем наличие загруженного второго файла напрямую через session_state
                                # Это позволяет определить загружен ли файл, даже если он ещё не обработан
                                has_categories_file = st.session_state.get('upload_categories_file') is not None
                                
                                # Проверяем наличие всех необходимых данных для таблицы 7
                                # Если файл загружен, пытаемся использовать обработанные данные
                                has_categories_data = (
                                    has_categories_file and
                                    'df_categories' in st.session_state and st.session_state.df_categories is not None and
                                    'categories_list' in st.session_state and st.session_state.categories_list is not None and
                                    'group_col_name' in st.session_state and st.session_state.group_col_name is not None and
                                    'year_month_col_name' in st.session_state and 'client_code_col_name' in st.session_state
                                )
                                
                                if has_categories_data:
                                    
                                    df_categories = st.session_state.df_categories
                                    categories = st.session_state.categories_list
                                    group_col = st.session_state.group_col_name
                                    year_month_col_cat = st.session_state.get('year_month_col_name', None)
                                    client_code_col = st.session_state.get('client_code_col_name', None)
                                    
                                    # Проверяем, что есть категории для обработки
                                    if not categories or len(categories) == 0:
                                        # Пропускаем создание таблицы 6, если нет категорий
                                        pass
                                    else:
                                        start_row_cohorts = 0
                                        worksheet_cohorts = None
                                        
                                        # Добавляем данные с листа 6 (если есть)
                                        if 'category_summary_table' in st.session_state and st.session_state.category_summary_table is not None:
                                            summary_table_excel = st.session_state.category_summary_table.copy()
                                            summary_table_excel.index.name = 'Метрика / Когорта'
                                            summary_table_excel.to_excel(writer, sheet_name="6. Присутствие когорты в других категориях", startrow=start_row_cohorts, index=True)
                                            worksheet_cohorts = writer.sheets["6. Присутствие когорты в других категориях"]
                                            
                                            # Форматируем верхнюю таблицу
                                            for row_idx in range(start_row_cohorts + 2, start_row_cohorts + len(summary_table_excel.index) + 2):
                                                for col_idx in range(2, len(summary_table_excel.columns) + 2):
                                                    cell = worksheet_cohorts.cell(row=row_idx, column=col_idx)
                                                    cell.alignment = ExcelAlignment(horizontal="center", vertical="center")
                                                    row_name = summary_table_excel.index[row_idx - start_row_cohorts - 2]
                                                    
                                                    if cell.value is not None and not isinstance(cell.value, str):
                                                        if row_name == 'Доля оттока из сети от когорты':
                                                            cell.value = float(cell.value) / 100.0
                                                            cell.number_format = '0.0%'
                                                        else:
                                                            cell.number_format = '0'
                                            
                                            # Форматируем заголовок строки
                                            for row_idx in range(start_row_cohorts + 2, start_row_cohorts + len(summary_table_excel.index) + 2):
                                                cell = worksheet_cohorts.cell(row=row_idx, column=1)
                                                cell.alignment = ExcelAlignment(horizontal="left", vertical="center")
                                            
                                            start_row_cohorts = start_row_cohorts + len(summary_table_excel.index) + 3
                                        
                                        if 'category_cohort_table' in st.session_state and st.session_state.category_cohort_table is not None:
                                            category_table_excel = st.session_state.category_cohort_table.copy()
                                            category_table_excel.index.name = 'Категория / Когорта'
                                            
                                            if worksheet_cohorts is None:
                                                category_table_excel.to_excel(writer, sheet_name="6. Присутствие когорты в других категориях", startrow=start_row_cohorts, index=True)
                                                worksheet_cohorts = writer.sheets["6. Присутствие когорты в других категориях"]
                                            else:
                                                category_table_excel.to_excel(writer, sheet_name="6. Присутствие когорты в других категориях", startrow=start_row_cohorts, index=True)
                                            
                                            # Форматируем таблицу с категориями
                                            for row_idx in range(start_row_cohorts + 2, start_row_cohorts + len(category_table_excel.index) + 2):
                                                for col_idx in range(2, len(category_table_excel.columns) + 2):
                                                    cell = worksheet_cohorts.cell(row=row_idx, column=col_idx)
                                                    cell.alignment = ExcelAlignment(horizontal="center", vertical="center")
                                                    if cell.value is not None and not isinstance(cell.value, str):
                                                        cell.number_format = '0'
                                            
                                            # Форматируем заголовок строки
                                            for row_idx in range(start_row_cohorts + 2, start_row_cohorts + len(category_table_excel.index) + 2):
                                                cell = worksheet_cohorts.cell(row=row_idx, column=1)
                                                cell.alignment = ExcelAlignment(horizontal="left", vertical="center")
                                            
                                            start_row_cohorts = start_row_cohorts + len(category_table_excel.index) + 3
                                        
                                        # Для каждой когорты создаем таблицу
                                        for cohort_idx, selected_cohort in enumerate(sorted_periods):
                                            # Определяем периоды начиная с выбранной когорты
                                            cohort_index = sorted_periods.index(selected_cohort) if selected_cohort in sorted_periods else 0
                                            periods_from_cohort = sorted_periods[cohort_index:]
                                            # Периоды ПОСЛЕ когорты (исключая период когорты) - для столбцов таблицы
                                            periods_after_cohort = periods_from_cohort[1:] if len(periods_from_cohort) > 1 else []
                                            
                                            # Пропускаем когорту, если нет периодов после неё
                                            if len(periods_after_cohort) == 0:
                                                continue
                                            
                                            # Получаем клиентов оттока для выбранной когорты
                                            period_clients_cache = st.session_state.get('period_clients_cache', None)
                                            client_cohorts_cache = st.session_state.get('client_cohorts_cache', None)
                                            churn_clients_set = set(get_churn_clients(df, year_month_col, client_col, sorted_periods, selected_cohort, period_clients_cache, client_cohorts_cache))
                                            churn_clients_set = {str(client) for client in churn_clients_set}
                                            
                                            # Создаем таблицу: категории по строкам, периоды ПОСЛЕ когорты по столбцам
                                            category_period_table = pd.DataFrame(index=categories, columns=periods_after_cohort)
                                            
                                            # Словари для итогов
                                            period_unique_clients = {period: set() for period in periods_after_cohort}
                                            category_unique_clients = {category: set() for category in categories}
                                            
                                            # Если есть столбец периода (Год-месяц / Год-неделя), фильтруем по периодам
                                            if year_month_col_cat is not None:
                                                # Нормализация периодов для сравнения (строки), чтобы избежать расхождения типов
                                                cat_period_str = df_categories[year_month_col_cat].astype(str).str.strip()
                                                for period in periods_after_cohort:
                                                    period_str = str(period).strip()
                                                    period_data = df_categories[cat_period_str == period_str]
                                                    
                                                    for category in categories:
                                                        category_period_data = period_data[
                                                            (period_data[group_col] == category) & 
                                                            (period_data[client_code_col].notna())
                                                        ]
                                                        
                                                        category_period_clients = set(
                                                            category_period_data[client_code_col].dropna().astype(str).unique()
                                                        )
                                                        
                                                        intersection = churn_clients_set & category_period_clients
                                                        category_period_table.loc[category, period] = len(intersection)
                                                        
                                                        period_unique_clients[period].update(intersection)
                                                        category_unique_clients[category].update(intersection)
                                            else:
                                                category_clients_dict = {}
                                                for category in categories:
                                                    category_data = df_categories[df_categories[group_col] == category]
                                                    client_codes = set(category_data[client_code_col].dropna().astype(str).unique())
                                                    category_clients_dict[category] = client_codes
                                                
                                                for period in periods_after_cohort:
                                                    for category in categories:
                                                        category_clients_set = category_clients_dict.get(category, set())
                                                        intersection = churn_clients_set & category_clients_set
                                                        category_period_table.loc[category, period] = len(intersection)
                                                        
                                                        period_unique_clients[period].update(intersection)
                                                        category_unique_clients[category].update(intersection)
                                            
                                            # Заполняем NaN нулями
                                            category_period_table = category_period_table.fillna(0).astype(int)
                                            
                                            # Создаем итоговую строку и столбец
                                            totals_row = pd.Series(
                                                {period: len(period_unique_clients[period]) for period in periods_after_cohort},
                                                name='Итого клиентов'
                                            )
                                            
                                            totals_col = pd.Series(
                                                {category: len(category_unique_clients[category]) for category in categories},
                                                name='Итого'
                                            )
                                            
                                            # Добавляем итоги в таблицу
                                            category_period_table_with_totals = category_period_table.copy()
                                            category_period_table_with_totals.loc['Итого клиентов'] = totals_row
                                            category_period_table_with_totals['Итого'] = totals_col
                                            
                                            # Вычисляем значение для ячейки пересечения
                                            all_category_clients = set()
                                            if year_month_col_cat is not None:
                                                periods_after_set = {str(p).strip() for p in periods_after_cohort}
                                                for category in categories:
                                                    category_data = df_categories[df_categories[group_col] == category]
                                                    period_col_str = category_data[year_month_col_cat].astype(str).str.strip()
                                                    category_data_filtered = category_data[period_col_str.isin(periods_after_set)]
                                                    category_clients = set(category_data_filtered[client_code_col].dropna().astype(str).unique())
                                                    all_category_clients.update(category_clients)
                                            else:
                                                for category in categories:
                                                    category_data = df_categories[df_categories[group_col] == category]
                                                    category_clients = set(category_data[client_code_col].dropna().astype(str).unique())
                                                    all_category_clients.update(category_clients)
                                            
                                            present_in_categories = churn_clients_set & all_category_clients
                                            category_period_table_with_totals.loc['Итого клиентов', 'Итого'] = len(present_in_categories)
                                            
                                            # Переупорядочиваем строки и столбцы
                                            new_index = ['Итого клиентов'] + [cat for cat in categories]
                                            category_period_table_with_totals = category_period_table_with_totals.reindex(new_index)
                                            new_columns = ['Итого'] + list(periods_after_cohort)
                                            category_period_table_with_totals = category_period_table_with_totals[new_columns]
                                            
                                            # Добавляем заголовок когорты
                                            if worksheet_cohorts is None:
                                                # Создаем новый лист
                                                category_period_table_with_totals.to_excel(
                                                    writer, 
                                                    sheet_name="6. Присутствие когорты в других категориях", 
                                                    startrow=start_row_cohorts, 
                                                    index=True
                                                )
                                                worksheet_cohorts = writer.sheets["6. Присутствие когорты в других категориях"]
                                                # Добавляем заголовок когорты
                                                last_col_letter = get_column_letter(len(new_columns) + 1)
                                                worksheet_cohorts.cell(row=start_row_cohorts + 1, column=1, value=f"Когорта: {selected_cohort}")
                                                worksheet_cohorts.merge_cells(f'A{start_row_cohorts + 1}:{last_col_letter}{start_row_cohorts + 1}')
                                                header_cell = worksheet_cohorts.cell(row=start_row_cohorts + 1, column=1)
                                                header_cell.font = Font(bold=True, size=12)
                                                header_cell.alignment = ExcelAlignment(horizontal="center", vertical="center")
                                                start_row_cohorts += 2
                                            else:
                                                # Добавляем заголовок когорты
                                                last_col_letter = get_column_letter(len(new_columns) + 1)
                                                worksheet_cohorts.cell(row=start_row_cohorts + 1, column=1, value=f"Когорта: {selected_cohort}")
                                                worksheet_cohorts.merge_cells(f'A{start_row_cohorts + 1}:{last_col_letter}{start_row_cohorts + 1}')
                                                header_cell = worksheet_cohorts.cell(row=start_row_cohorts + 1, column=1)
                                                header_cell.font = Font(bold=True, size=12)
                                                header_cell.alignment = ExcelAlignment(horizontal="center", vertical="center")
                                                start_row_cohorts += 2
                                                
                                                # Записываем таблицу на тот же лист
                                                category_period_table_with_totals.to_excel(
                                                    writer, 
                                                    sheet_name="6. Присутствие когорты в других категориях", 
                                                    startrow=start_row_cohorts, 
                                                    index=True
                                                )
                                            
                                            # Форматируем таблицу
                                            for row_idx in range(start_row_cohorts + 2, start_row_cohorts + len(category_period_table_with_totals.index) + 2):
                                                for col_idx in range(2, len(category_period_table_with_totals.columns) + 2):
                                                    cell = worksheet_cohorts.cell(row=row_idx, column=col_idx)
                                                    cell.alignment = ExcelAlignment(horizontal="center", vertical="center")
                                                    if cell.value is not None and not isinstance(cell.value, str):
                                                        cell.number_format = '0'
                                            
                                            # Форматируем заголовок строки
                                            for row_idx in range(start_row_cohorts + 2, start_row_cohorts + len(category_period_table_with_totals.index) + 2):
                                                cell = worksheet_cohorts.cell(row=row_idx, column=1)
                                                cell.alignment = ExcelAlignment(horizontal="left", vertical="center")
                                            
                                            # Обновляем начальную строку для следующей таблицы (таблица + 2 пустые строки)
                                            start_row_cohorts = start_row_cohorts + len(category_period_table_with_totals.index) + 3
                                
                                # Таблица 7: Сводная таблица по всем когортам
                                # Таблица 8 всегда создаётся с базовыми метриками (1-5)
                                # Метрики 6-9 добавляются только при наличии данных категорий
                                if st.session_state.get('churn_table') is not None:
                                    churn_table = st.session_state.churn_table
                                    has_categories_file_excel = (
                                        st.session_state.get('upload_categories_file') is not None or
                                        st.session_state.get('category_summary_table') is not None
                                    )
                                    summary_data = {}
                                    
                                    # 1–3. Базовые метрики (всегда)
                                    summary_data['Кол-во клиентов в когорте'] = {}
                                    for _, row in churn_table.iterrows():
                                        cohort = row['Когорта']
                                        summary_data['Кол-во клиентов в когорте'][cohort] = int(row['Кол-во клиентов когорты'])
                                    summary_data['Накопительное кол-во вернувшихся в категорию'] = {}
                                    for _, row in churn_table.iterrows():
                                        cohort = row['Когорта']
                                        summary_data['Накопительное кол-во вернувшихся в категорию'][cohort] = _churn_int(row['Накопительное кол-во возврата'])
                                    summary_data['Накопительное кол-во вернувшихся в категорию %'] = {}
                                    for _, row in churn_table.iterrows():
                                        cohort = row['Когорта']
                                        v_ret = row['Накопительный % возврата']
                                        summary_data['Накопительное кол-во вернувшихся в категорию %'][cohort] = v_ret if v_ret == '-' else f"{float(v_ret):.1f}%"
                                    
                                    # 4–5. Отток из категории (из первого файла — всегда)
                                    summary_data['Отток из категории когорты'] = {}
                                    for _, row in churn_table.iterrows():
                                        cohort = row['Когорта']
                                        summary_data['Отток из категории когорты'][cohort] = _churn_int(row['Отток кол-во'])
                                    summary_data['Отток из категории когорты %'] = {}
                                    for _, row in churn_table.iterrows():
                                        cohort = row['Когорта']
                                        v = row['Отток %']
                                        summary_data['Отток из категории когорты %'][cohort] = v if v == '-' else f"{float(v):.1f}%"
                                    
                                    if has_categories_file_excel:
                                        _pa_ex = st.session_state.get('period_after_label', 'месяца')
                                        _k_ит = f"Итого присутствуют в других категориях после {_pa_ex} когорты"
                                        _k_доля = f"Доля присутствуют в других категориях после {_pa_ex} когорты"
                                        _k_кол = f"Кол-во клиентов когорты в других категориях после {_pa_ex} когорты"
                                        _k_кол_pct = f"Кол-во клиентов когорты в других категориях после {_pa_ex} когорты %"
                                        summary_data[_k_кол] = {}
                                        summary_data[_k_кол_pct] = {}
                                        summary_data['Отток из сети'] = {}
                                        summary_data['Отток из сети %'] = {}
                                        for cohort in sorted_periods:
                                            summary_data[_k_кол][cohort] = 0
                                            summary_data[_k_кол_pct][cohort] = 0.0
                                            summary_data['Отток из сети'][cohort] = 0
                                            summary_data['Отток из сети %'][cohort] = 0.0
                                        
                                        if 'category_summary_table' in st.session_state and st.session_state.category_summary_table is not None:
                                            category_summary = st.session_state.category_summary_table
                                            if _k_ит in category_summary.index:
                                                for cohort in sorted_periods:
                                                    if cohort in category_summary.columns:
                                                        value = category_summary.loc[_k_ит, cohort]
                                                        summary_data[_k_кол][cohort] = int(value) if pd.notna(value) else 0
                                            if _k_доля in category_summary.index:
                                                for cohort in sorted_periods:
                                                    if cohort in category_summary.columns:
                                                        value = category_summary.loc[_k_доля, cohort]
                                                        if pd.notna(value):
                                                            summary_data[_k_кол_pct][cohort] = value
                                            else:
                                                for cohort in sorted_periods:
                                                    cohort_size = summary_data['Кол-во клиентов в когорте'].get(cohort, 0)
                                                    present_after_count = summary_data[_k_кол].get(cohort, 0)
                                                    if cohort_size > 0:
                                                        percent = (present_after_count / cohort_size) * 100
                                                        summary_data[_k_кол_pct][cohort] = percent
                                            if 'Отток из сети' in category_summary.index:
                                                for cohort in sorted_periods:
                                                    if cohort in category_summary.columns:
                                                        value = category_summary.loc['Отток из сети', cohort]
                                                        summary_data['Отток из сети'][cohort] = int(value) if pd.notna(value) else 0
                                            if 'Доля оттока из сети от когорты' in category_summary.index:
                                                for cohort in sorted_periods:
                                                    if cohort in category_summary.columns:
                                                        value = category_summary.loc['Доля оттока из сети от когорты', cohort]
                                                        if pd.notna(value):
                                                            summary_data['Отток из сети %'][cohort] = value
                                    
                                    # Создаем DataFrame
                                    summary_df = pd.DataFrame(summary_data, index=sorted_periods).T
                                    summary_df.index.name = 'Метрика / Когорта'
                                    
                                    # Записываем в Excel
                                    summary_df.to_excel(writer, sheet_name="7. Сводная таблица по всем когортам", startrow=0, index=True)
                                    worksheet_summary = writer.sheets["7. Сводная таблица по всем когортам"]
                                    
                                    # Форматируем таблицу
                                    for row_idx in range(2, len(summary_df.index) + 2):
                                        for col_idx in range(2, len(summary_df.columns) + 2):
                                            cell = worksheet_summary.cell(row=row_idx, column=col_idx)
                                            cell.alignment = ExcelAlignment(horizontal="center", vertical="center")
                                            row_name = summary_df.index[row_idx - 2]
                                            
                                            if cell.value is not None and not isinstance(cell.value, str):
                                                if '%' in row_name:
                                                    # Процентные колонки
                                                    cell.value = float(cell.value) / 100.0 if isinstance(cell.value, (int, float)) and cell.value > 1 else float(cell.value)
                                                    cell.number_format = '0.0%'
                                                else:
                                                    # Числовые колонки
                                                    cell.number_format = '0'
                                    
                                    # Форматируем заголовок строки
                                    for row_idx in range(2, len(summary_df.index) + 2):
                                        cell = worksheet_summary.cell(row=row_idx, column=1)
                                        cell.alignment = ExcelAlignment(horizontal="left", vertical="center")
                                
                                # Удаляем пустой лист по умолчанию
                                if 'Sheet' in workbook.sheetnames:
                                    workbook.remove(workbook['Sheet'])
                            
                            buffer.seek(0)
                            return buffer.getvalue()
                        
                        # CSS для увеличения размера кнопок загрузки
                        st.markdown("""
                        <style>
                        div[data-testid="stDownloadButton"] > button {
                            height: 60px !important;
                            font-size: 20px !important;
                            font-weight: bold !important;
                            padding: 15px 30px !important;
                        }
                        div[data-testid="stDownloadButton"] > button > div > p {
                            font-size: 20px !important;
                            font-weight: bold !important;
                        }
                        </style>
                        """, unsafe_allow_html=True)
                        
                        # Создаем колонки для горизонтального размещения кнопок
                        col_excel_button, col_pdf_button = st.columns(2)
                        
                        # Генерируем файл каждый раз при рендеринге (данные могут обновиться)
                        # Всегда генерируем отчет заново, чтобы включить все актуальные данные
                        try:
                            excel_data_full = create_full_report_excel()
                            # Сохраняем для возможного использования в будущем
                            st.session_state.excel_report_data = excel_data_full
                        except Exception as e:
                            # Если ошибка, используем сохраненный файл как fallback
                            if 'excel_report_data' in st.session_state and st.session_state.excel_report_data is not None:
                                excel_data_full = st.session_state.excel_report_data
                                st.warning(f"Использован сохраненный отчет. Ошибка при генерации: {str(e)}")
                            else:
                                st.error(f"Ошибка при генерации отчета: {str(e)}")
                                excel_data_full = b""  # Пустой файл
                        
                        # Имя файла с продуктом построения когорт
                        _df = st.session_state.get('df')
                        if _df is not None and len(_df.columns) > 0:
                            _pc = _df.columns[0]
                            _upl = sorted(_df[_pc].dropna().astype(str).str.strip().unique())
                            _upl = [p for p in _upl if p]
                            _suffix = "_".join(_upl)
                            _suffix = re.sub(r'[\\/:*?"<>|]', '_', _suffix)[:80].strip('._ ') if _suffix else ""
                        else:
                            _suffix = ""
                        _excel_name = f"полный_отчёт_когортный_анализ_{_suffix}_{info['first_period']}_{info['last_period']}.xlsx" if _suffix else f"полный_отчёт_когортный_анализ_{info['first_period']}_{info['last_period']}.xlsx"
                        with col_excel_button:
                            st.download_button(
                                label="📥 Скачать полный отчёт в Excel",
                                data=excel_data_full,
                                file_name=_excel_name,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key="download_full_report"
                            )
                        
                        # Создаем функцию для генерации аналитического PDF отчёта
                        def create_analysis_pdf():
                            """Создает PDF отчёт с графиками и анализом"""
                            buffer = io.BytesIO()
                            
                            # Регистрируем шрифт с поддержкой кириллицы
                            font_name = 'Helvetica'
                            font_name_bold = 'Helvetica-Bold'
                            
                            try:
                                # Пытаемся найти системный шрифт с поддержкой кириллицы
                                if platform.system() == 'Windows':
                                    # Пути к стандартным шрифтам Windows с поддержкой кириллицы
                                    windows_fonts = [
                                        r'C:\Windows\Fonts\arial.ttf',
                                        r'C:\Windows\Fonts\calibri.ttf',
                                        r'C:\Windows\Fonts\comic.ttf',
                                        r'C:\Windows\Fonts\cour.ttf',
                                    ]
                                    
                                    # Регистрируем первый доступный шрифт
                                    for font_path in windows_fonts:
                                        if os.path.exists(font_path):
                                            try:
                                                font_name = 'CyrillicFont'
                                                font_name_bold = 'CyrillicFont-Bold'
                                                pdfmetrics.registerFont(TTFont(font_name, font_path))
                                                pdfmetrics.registerFont(TTFont(font_name_bold, font_path))
                                                break
                                            except Exception as e:
                                                continue
                                elif platform.system() == 'Linux':
                                    # Пути к стандартным шрифтам Linux с поддержкой кириллицы
                                    linux_fonts = [
                                        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                                        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
                                        '/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf',
                                        '/usr/share/fonts/truetype/ttf-dejavu/DejaVuSans.ttf',
                                        '/usr/share/fonts/TTF/DejaVuSans.ttf',
                                    ]
                                    
                                    # Регистрируем первый доступный шрифт
                                    for font_path in linux_fonts:
                                        if os.path.exists(font_path):
                                            try:
                                                font_name = 'CyrillicFont'
                                                font_name_bold = 'CyrillicFont-Bold'
                                                pdfmetrics.registerFont(TTFont(font_name, font_path))
                                                pdfmetrics.registerFont(TTFont(font_name_bold, font_path))
                                                break
                                            except Exception as e:
                                                continue
                            except Exception as e:
                                pass  # Используем стандартные шрифты в случае ошибки
                            
                            # Получаем данные из session state
                            cohort_matrix = st.session_state.cohort_matrix
                            sorted_periods = st.session_state.sorted_periods
                            accumulation_matrix = st.session_state.accumulation_matrix
                            accumulation_percent_matrix = st.session_state.accumulation_percent_matrix
                            inflow_matrix = st.session_state.inflow_matrix
                            churn_table = st.session_state.churn_table
                            
                            # Создаем PDF документ
                            doc = SimpleDocTemplate(buffer, pagesize=A4)
                            story = []
                            styles = getSampleStyleSheet()
                            
                            # Стили с поддержкой кириллицы
                            title_style = ParagraphStyle(
                                'CustomTitle',
                                parent=styles['Heading1'],
                                fontName=font_name_bold,
                                fontSize=24,
                                textColor=colors.HexColor('#1f77b4'),
                                spaceAfter=30,
                                alignment=TA_CENTER
                            )
                            
                            heading_style = ParagraphStyle(
                                'CustomHeading',
                                parent=styles['Heading2'],
                                fontName=font_name_bold,
                                fontSize=16,
                                textColor=colors.HexColor('#1f77b4'),
                                spaceAfter=12,
                                spaceBefore=12
                            )
                            
                            # Стиль для обычного текста с поддержкой кириллицы
                            normal_style = ParagraphStyle(
                                'CustomNormal',
                                parent=styles['Normal'],
                                fontName=font_name,
                                fontSize=10
                            )
                            
                            # Стиль для заголовков третьего уровня с поддержкой кириллицы
                            heading3_style = ParagraphStyle(
                                'CustomHeading3',
                                parent=styles['Heading3'],
                                fontName=font_name_bold,
                                fontSize=12,
                                textColor=colors.HexColor('#1f77b4'),
                                spaceAfter=8,
                                spaceBefore=8
                            )
                            
                            # Титульная страница
                            story.append(Paragraph("КОГОРТНЫЙ АНАЛИЗ", title_style))
                            story.append(Spacer(1, 0.3*inch))
                            story.append(Paragraph(f"Период анализа: {info['first_period']} - {info['last_period']}", normal_style))
                            story.append(Paragraph(f"Количество когорт: {info['num_periods']}", normal_style))
                            story.append(Paragraph(f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}", normal_style))
                            story.append(PageBreak())
                            
                            # Раздел 1: Общая статистика
                            story.append(Paragraph("1. ОБЩАЯ СТАТИСТИКА", heading_style))
                            
                            # Диагональные значения (размер когорт)
                            diagonal_values = {period: cohort_matrix.loc[period, period] for period in sorted_periods}
                            
                            stats_data = [
                                ['Метрика', 'Значение'],
                                ['Всего когорт', str(info['num_periods'])],
                                ['Период начала', info['first_period']],
                                ['Период окончания', info['last_period']],
                                ['Максимальный размер когорты', f"{int(info['max_clients'])} ({info['max_period']})"],
                                ['Минимальный размер когорты', f"{int(info['min_clients'])} ({info['min_period']})"],
                                ['Средний размер когорты', f"{int(np.mean(list(diagonal_values.values())))}"],
                                ['Общее количество уникальных клиентов', f"{int(sum(diagonal_values.values()))}"]
                            ]
                            
                            stats_table = Table(stats_data, colWidths=[4*inch, 3*inch])
                            stats_table.setStyle(TableStyle([
                                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
                                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                                ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
                                ('FONTNAME', (0, 1), (-1, -1), font_name),
                                ('FONTSIZE', (0, 0), (-1, 0), 12),
                                ('FONTSIZE', (0, 1), (-1, -1), 10),
                                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                                ('GRID', (0, 0), (-1, -1), 1, colors.black)
                            ]))
                            story.append(stats_table)
                            story.append(Spacer(1, 0.3*inch))
                            
                            # График 1: Динамика размера когорт
                            story.append(Paragraph("2. ДИНАМИКА РАЗМЕРА КОГОРТ", heading_style))
                            
                            fig, ax = plt.subplots(figsize=(10, 6))
                            cohort_sizes = [diagonal_values[p] for p in sorted_periods]
                            ax.plot(range(len(sorted_periods)), cohort_sizes, marker='o', linewidth=2, markersize=8, color='#1f77b4')
                            ax.set_xlabel('Период', fontsize=12, fontweight='bold')
                            ax.set_ylabel('Количество клиентов', fontsize=12, fontweight='bold')
                            ax.set_title('Динамика размера когорт по периодам', fontsize=14, fontweight='bold', pad=20)
                            ax.set_xticks(range(len(sorted_periods)))
                            ax.set_xticklabels(sorted_periods, rotation=45, ha='right')
                            ax.grid(True, alpha=0.3)
                            ax.set_facecolor('#f8f9fa')
                            
                            for i, (period, size) in enumerate(zip(sorted_periods, cohort_sizes)):
                                ax.annotate(f'{int(size)}', (i, size), textcoords="offset points", xytext=(0,10), ha='center', fontsize=9)
                            
                            plt.tight_layout()
                            img_buffer1 = io.BytesIO()
                            plt.savefig(img_buffer1, format='png', dpi=150, bbox_inches='tight')
                            img_buffer1.seek(0)
                            plt.close()
                            
                            img1 = Image(img_buffer1, width=6*inch, height=3.6*inch)
                            story.append(img1)
                            story.append(Spacer(1, 0.3*inch))
                            
                            # График 2: Тепловая карта возврата в %
                            story.append(Paragraph("3. ТЕПЛОВАЯ КАРТА ВОЗВРАТА В %", heading_style))
                            
                            # Создаём упрощённую матрицу для визуализации (первые 15 когорт и периодов)
                            max_cohorts = min(15, len(sorted_periods))
                            matrix_vis = accumulation_percent_matrix.iloc[:max_cohorts, :max_cohorts]
                            
                            fig, ax = plt.subplots(figsize=(12, 10))
                            sns.heatmap(matrix_vis, annot=True, fmt='.1f', cmap='RdYlGn', 
                                       cbar_kws={'label': 'Процент возврата (%)'}, 
                                       ax=ax, vmin=0, vmax=100, linewidths=0.5, linecolor='gray')
                            ax.set_title('Тепловая карта накопления возврата клиентов (%)', fontsize=14, fontweight='bold', pad=20)
                            ax.set_xlabel('Период', fontsize=12, fontweight='bold')
                            ax.set_ylabel('Когорта', fontsize=12, fontweight='bold')
                            
                            plt.tight_layout()
                            img_buffer2 = io.BytesIO()
                            plt.savefig(img_buffer2, format='png', dpi=150, bbox_inches='tight')
                            img_buffer2.seek(0)
                            plt.close()
                            
                            img2 = Image(img_buffer2, width=6*inch, height=5*inch)
                            story.append(img2)
                            story.append(Spacer(1, 0.3*inch))
                            
                            # График 3: Отток по когортам
                            story.append(Paragraph("4. АНАЛИЗ ОТТОКА КЛИЕНТОВ", heading_style))
                            
                            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
                            
                            # Столбчатая диаграмма оттока в количестве (последняя когорта с "-" даёт 0)
                            churn_counts_series = pd.to_numeric(churn_table['Отток кол-во'], errors='coerce').fillna(0).astype(int)
                            churn_counts = churn_counts_series.values[:15]
                            cohorts_display = churn_table['Когорта'].values[:15]
                            churn_mean = churn_counts_series.replace(0, np.nan).mean()
                            if pd.isna(churn_mean):
                                churn_mean = 0
                            colors_churn = ['#d62728' if x > churn_mean else '#ff7f0e' for x in churn_counts]
                            ax1.barh(range(len(cohorts_display)), churn_counts, color=colors_churn)
                            ax1.set_yticks(range(len(cohorts_display)))
                            ax1.set_yticklabels(cohorts_display, fontsize=9)
                            ax1.set_xlabel('Количество клиентов оттока', fontsize=11, fontweight='bold')
                            ax1.set_title('Отток клиентов из категории по когортам', fontsize=12, fontweight='bold')
                            ax1.grid(True, alpha=0.3, axis='x')
                            
                            # Столбчатая диаграмма оттока в процентах
                            churn_percents_series = pd.to_numeric(churn_table['Отток %'], errors='coerce').fillna(0)
                            churn_percents = churn_percents_series.values[:15]
                            churn_pct_mean = churn_percents_series.replace(0, np.nan).mean()
                            if pd.isna(churn_pct_mean):
                                churn_pct_mean = 0
                            colors_churn_pct = ['#d62728' if x > churn_pct_mean else '#ff7f0e' for x in churn_percents]
                            ax2.barh(range(len(cohorts_display)), churn_percents, color=colors_churn_pct)
                            ax2.set_yticks(range(len(cohorts_display)))
                            ax2.set_yticklabels(cohorts_display, fontsize=9)
                            ax2.set_xlabel('Процент оттока (%)', fontsize=11, fontweight='bold')
                            ax2.set_title('Процент оттока по когортам', fontsize=12, fontweight='bold')
                            ax2.grid(True, alpha=0.3, axis='x')
                            
                            plt.tight_layout()
                            img_buffer4 = io.BytesIO()
                            plt.savefig(img_buffer4, format='png', dpi=150, bbox_inches='tight')
                            img_buffer4.seek(0)
                            plt.close()
                            
                            img4 = Image(img_buffer4, width=7*inch, height=3.6*inch)
                            story.append(img4)
                            story.append(Spacer(1, 0.3*inch))
                            
                            # Таблицы с ключевыми метриками
                            story.append(Paragraph("5. КЛЮЧЕВЫЕ МЕТРИКИ", heading_style))
                            
                            # Топ-5 когорт по размеру
                            story.append(Paragraph("Топ-5 когорт по размеру:", heading3_style))
                            top5_size = sorted(diagonal_values.items(), key=lambda x: x[1], reverse=True)[:5]
                            top5_data = [['Место', 'Когорта', 'Количество клиентов']]
                            for i, (period, size) in enumerate(top5_size, 1):
                                top5_data.append([str(i), period, str(int(size))])
                            
                            top5_table = Table(top5_data, colWidths=[0.8*inch, 2.5*inch, 2*inch])
                            top5_table.setStyle(TableStyle([
                                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
                                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
                                ('FONTNAME', (0, 1), (-1, -1), font_name),
                                ('FONTSIZE', (0, 0), (-1, 0), 10),
                                ('FONTSIZE', (0, 1), (-1, -1), 10),
                                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                                ('GRID', (0, 0), (-1, -1), 1, colors.black)
                            ]))
                            story.append(top5_table)
                            story.append(Spacer(1, 0.2*inch))
                            
                            # Топ-5 когорт по проценту возврата
                            story.append(Paragraph("Топ-5 когорт по проценту возврата:", heading3_style))
                            churn_return_numeric = pd.to_numeric(churn_table['Накопительный % возврата'], errors='coerce')
                            churn_sorted_return = churn_table.assign(_pct=churn_return_numeric).dropna(subset=['_pct']).sort_values('_pct', ascending=False).drop(columns=['_pct'])
                            top5_return_data = [['Место', 'Когорта', 'Процент возврата', 'Размер когорты']]
                            for i, row in enumerate(churn_sorted_return.head(5).itertuples(index=False), 1):
                                top5_return_data.append([
                                    str(i), 
                                    row[0],  # Когорта
                                    f"{row[3]:.1f}%",  # Накопительный % возврата
                                    str(int(row[1]))  # Кол-во клиентов когорты
                                ])
                            
                            top5_return_table = Table(top5_return_data, colWidths=[0.8*inch, 2*inch, 1.5*inch, 1.5*inch])
                            top5_return_table.setStyle(TableStyle([
                                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ca02c')),
                                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
                                ('FONTNAME', (0, 1), (-1, -1), font_name),
                                ('FONTSIZE', (0, 0), (-1, 0), 10),
                                ('FONTSIZE', (0, 1), (-1, -1), 10),
                                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                                ('GRID', (0, 0), (-1, -1), 1, colors.black)
                            ]))
                            story.append(top5_return_table)
                            story.append(Spacer(1, 0.2*inch))
                            
                            # Когорты с максимальным оттоком (исключаем последнюю когорту с "-")
                            story.append(Paragraph("Топ-5 когорт с наибольшим оттоком:", heading3_style))
                            churn_churn_numeric = pd.to_numeric(churn_table['Отток %'], errors='coerce')
                            churn_sorted_churn = churn_table.assign(_pct=churn_churn_numeric).dropna(subset=['_pct']).sort_values('_pct', ascending=False).drop(columns=['_pct'])
                            top5_churn_data = [['Место', 'Когорта', 'Отток (%)', 'Отток (кол-во)']]
                            for i, row in enumerate(churn_sorted_churn.head(5).itertuples(index=False), 1):
                                top5_churn_data.append([
                                    str(i),
                                    row[0],  # Когорта
                                    f"{float(row[5]):.1f}%",  # Отток %
                                    str(_churn_int(row[4]))  # Отток кол-во
                                ])
                            
                            top5_churn_table = Table(top5_churn_data, colWidths=[0.8*inch, 2*inch, 1.5*inch, 1.5*inch])
                            top5_churn_table.setStyle(TableStyle([
                                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d62728')),
                                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
                                ('FONTNAME', (0, 1), (-1, -1), font_name),
                                ('FONTSIZE', (0, 0), (-1, 0), 10),
                                ('FONTSIZE', (0, 1), (-1, -1), 10),
                                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                                ('GRID', (0, 0), (-1, -1), 1, colors.black)
                            ]))
                            story.append(top5_churn_table)
                            story.append(Spacer(1, 0.3*inch))
                            
                            # Выводы и рекомендации
                            story.append(Paragraph("6. ВЫВОДЫ И РЕКОМЕНДАЦИИ", heading_style))
                            
                            avg_return = pd.to_numeric(churn_table['Накопительный % возврата'], errors='coerce').mean()
                            avg_churn = pd.to_numeric(churn_table['Отток %'], errors='coerce').mean()
                            if pd.isna(avg_return):
                                avg_return = 0
                            if pd.isna(avg_churn):
                                avg_churn = 0
                            
                            top5_size = sorted(diagonal_values.items(), key=lambda x: x[1], reverse=True)[:5]
                            conclusions = [
                                f"• Средний процент возврата клиентов: {avg_return:.1f}%",
                                f"• Средний процент оттока: {avg_churn:.1f}%",
                                f"• Наиболее стабильная когорта (по размеру): {top5_size[0][0]} ({int(top5_size[0][1])} клиентов)",
                                f"• Когорта с наилучшим возвратом: {churn_sorted_return.iloc[0, 0]} ({churn_sorted_return.iloc[0, 3]:.1f}%)",
                                f"• Когорта с наибольшим оттоком требует внимания: {churn_sorted_churn.iloc[0, 0]} ({float(churn_sorted_churn.iloc[0, 5]):.1f}%)"
                            ]
                            
                            for conclusion in conclusions:
                                story.append(Paragraph(conclusion, normal_style))
                                story.append(Spacer(1, 0.1*inch))
                            
                            # Собираем PDF
                            doc.build(story)
                            buffer.seek(0)
                            return buffer.getvalue()
                        
                        # Генерируем PDF при нажатии кнопки
                        pdf_data = create_analysis_pdf()
                        
                        with col_pdf_button:
                            st.download_button(
                                label="📊 Скачать анализ отчёта в PDF",
                                data=pdf_data,
                                file_name=f"анализ_когортный_{info['first_period']}_{info['last_period']}.pdf",
                                mime="application/pdf",
                                use_container_width=True,
                                key="download_analysis_pdf"
                            )
                        # Продукт построения когорт — уникальные значения из первого столбца первого документа
                        df_first = st.session_state.get('df')
                        if df_first is not None and len(df_first.columns) > 0:
                            product_col = df_first.columns[0]
                            unique_products = sorted(df_first[product_col].dropna().astype(str).str.strip().unique())
                            unique_products = [p for p in unique_products if p]
                            if unique_products:
                                products_text = ", ".join(unique_products)
                                st.markdown(f"""
                                <p style="font-size: 1.5rem; font-weight: 600; margin-top: 16px;">
                                    Продукт построения когорт: <span style="color: #0d6efd; font-weight: 700;">{products_text}</span>
                                </p>
                                """, unsafe_allow_html=True)
                else:
                    st.info("⏳ Загрузите файл и дождитесь завершения расчётов для генерации отчётов")
                
                # Отображение матрицы (только если данные готовы)
                if info:
                    # Уменьшаем отступ перед блоком матриц
                    st.markdown("<div style='margin-top: 5px;'></div>", unsafe_allow_html=True)
                    
                    # Добавляем CSS для компактного отображения таблицы без прокрутки
                    st.markdown("""
                    <style>
                    div[data-testid="stDataFrame"] > div {
                        overflow: visible !important;
                    }
                    div[data-testid="stDataFrame"] table {
                        font-size: 0.7rem !important;
                        width: 100% !important;
                    }
                    /* Убираем overflow с внутренних контейнеров таблицы */
                    div[data-testid="stDataFrame"] > div > div {
                        overflow: visible !important;
                    }
                    div[data-testid="stDataFrame"] th, 
                    div[data-testid="stDataFrame"] td {
                        padding: 0.2rem 0.4rem !important;
                        font-size: 0.7rem !important;
                        white-space: nowrap !important;
                        text-align: center !important;
                    }
                    div[data-testid="stDataFrame"] table th,
                    div[data-testid="stDataFrame"] table td {
                        text-align: center !important;
                    }
                    </style>
                    """, unsafe_allow_html=True)
                    
                    # Объединенный блок с переключателем отображения
                    # CSS стили для красивого оформления блока
                    st.markdown("""
                    <style>
                    /* Стили для блока с таблицей */
                    .matrix-block-container {
                        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                        padding: 20px;
                        border-radius: 15px;
                        box-shadow: 0 10px 30px rgba(0,0,0,0.2);
                        margin-bottom: 20px;
                    }
                    
                    /* Стили для кнопок переключения */
                    .stRadio > div {
                        background: transparent;
                        padding: 0;
                        border-radius: 0;
                        box-shadow: none;
                        border: none;
                        display: flex;
                        flex-direction: row;
                        gap: 10px;
                        align-items: stretch;
                    }
                    
                    .stRadio > div > label {
                        background: white !important;
                        color: #333 !important;
                        padding: 12px 8px !important;
                        border-radius: 8px !important;
                        margin: 0 !important;
                        font-weight: 700 !important;
                        font-size: 0.75rem !important;
                        line-height: 1.2 !important;
                        transition: all 0.3s ease !important;
                        border: 2px solid #ccc !important;
                        box-shadow: 0 2px 4px rgba(0,0,0,0.1) !important;
                        cursor: pointer !important;
                        text-align: center !important;
                        flex: 1 1 0 !important;
                        min-width: 0 !important;
                        min-height: 50px !important;
                        height: auto !important;
                        max-height: 60px !important;
                        display: flex !important;
                        align-items: center !important;
                        justify-content: center !important;
                        white-space: normal !important;
                        word-wrap: break-word !important;
                        overflow: hidden !important;
                    }
                    
                    .stRadio > div > label:hover {
                        transform: translateY(-2px) !important;
                        box-shadow: 0 4px 8px rgba(0,0,0,0.15) !important;
                        background: #f5f5f5 !important;
                        border-color: #999 !important;
                    }
                    
                    .stRadio > div > label[data-baseweb="radio"]:has(input:checked) {
                        background: white !important;
                        color: #333 !important;
                        border-color: #666 !important;
                        box-shadow: 0 4px 10px rgba(0, 0, 0, 0.2) !important;
                    }
                    
                    .stRadio input[type="radio"]:checked + label {
                        background: white !important;
                        color: #333 !important;
                    }
                    
                    /* Стили для таблицы - только базовое оформление, не мешаем встроенному тулбару */
                    div[data-testid="stDataFrame"] {
                        background: white;
                        border-radius: 10px;
                        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
                        max-width: 100% !important;
                    }
                    
                    /* Стили для блока кодов клиентов */
                    .clients-block {
                        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                        padding: 15px;
                        border-radius: 10px;
                        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
                    }
                    
                    /* Стили для описания */
                    .description-block {
                        background: transparent;
                        padding: 15px;
                        border-radius: 10px;
                        margin-bottom: 15px;
                        box-shadow: none;
                        color: inherit;
                    }
                    
                    /* Контейнер для кнопок - ограничиваем ширину как у описания */
                    .stRadio {
                        max-width: 100%;
                    }
                    
                    /* Ограничиваем ширину контейнера кнопок и растягиваем на всю ширину */
                    div[data-testid="stRadio"] {
                        max-width: 100% !important;
                        width: 100% !important;
                    }
                    
                    /* Растягиваем контейнер с кнопками на всю ширину колонки */
                    .stRadio > div {
                        width: 100% !important;
                        display: flex !important;
                        flex-direction: row !important;
                        gap: 10px !important;
                    }
                    
                    /* Стили для кнопок Excel и PDF - светлые белые с большим шрифтом */
                    div[data-testid="stDownloadButton"] button,
                    div[data-testid="stButton"] button {
                        background: #f8f9fa !important;
                        color: #333 !important;
                        padding: 15px 20px !important;
                        border-radius: 8px !important;
                        margin: 0 !important;
                        font-weight: 700 !important;
                        font-size: 1.1rem !important;
                        line-height: 1.3 !important;
                        transition: all 0.3s ease !important;
                        border: 2px solid #e0e0e0 !important;
                        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.05) !important;
                        cursor: pointer !important;
                        text-align: center !important;
                        min-height: 60px !important;
                        height: auto !important;
                        display: flex !important;
                        align-items: center !important;
                        justify-content: center !important;
                        white-space: normal !important;
                        word-wrap: break-word !important;
                        width: 100% !important;
                    }
                    
                    div[data-testid="stDownloadButton"] button:hover,
                    div[data-testid="stButton"] button:hover {
                        transform: translateY(-2px) !important;
                        box-shadow: 0 4px 8px rgba(0, 0, 0, 0.1) !important;
                        background: #ffffff !important;
                        border-color: #d0d0d0 !important;
                    }
                    
                    div[data-testid="stDownloadButton"] button:active,
                    div[data-testid="stButton"] button:active {
                        transform: translateY(0) !important;
                        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.05) !important;
                    }
                    </style>
                    """, unsafe_allow_html=True)
                    
                    # Создаем колонки для выравнивания кнопок с блоком описания
                    # Кнопки занимают всю ширину до блока кодов клиентов (соотношение 4:1 как у таблицы)
                    col_buttons_container, col_empty = st.columns([4, 1])
                    
                    with col_buttons_container:
                        # Переключатель для выбора типа отображения (горизонтально, на уровне с таблицей)
                        view_type = st.radio(
                            "",
                            options=[
                                "Динамика уникальных клиентов когорт",
                                "Динамика накопления возврата",
                                "Динамика накопления возврата в %",
                                "Приток возврата в %",
                                "Отток клиентов из категории"
                            ],
                            horizontal=True,
                            key="view_type_selector"
                        )
                    
                    # Уменьшаем отступ между кнопками и таблицей
                    st.markdown("<div style='margin-top: 5px;'></div>", unsafe_allow_html=True)
                    
                    # Основной контент
                    # Инициализируем переменные для таблицы и описания
                    display_matrix = None
                    description_text = ""
                    view_key = ""
                    
                    # Подготовка данных в зависимости от выбранного типа
                    if view_type == "Динамика уникальных клиентов когорт":
                        # Применяем цветовое форматирование; нулевые значения скрываем
                        matrix_int = cohort_matrix.astype(int)
                        display_matrix = apply_matrix_color_gradient(matrix_int.astype(float), horizontal_dynamics=True, hide_before_diagonal=True, hide_zeros=True)
                        display_matrix = display_matrix.format(precision=0, thousands=',', decimal='.')
                        description_text = "Диагональ показывает количество уникальных клиентов в каждом периоде. Пересечения показывают количество клиентов, которые были активны в обоих периодах."
                        view_key = "cohort"
                        
                    elif view_type == "Динамика накопления возврата":
                        accumulation_matrix = st.session_state.accumulation_matrix
                        matrix_int_accum = accumulation_matrix.astype(int)
                        display_matrix = apply_matrix_color_gradient(matrix_int_accum.astype(float), hide_zeros=True)
                        display_matrix = display_matrix.format(precision=0, thousands=',', decimal='.')
                        description_text = "Показывает накопление уникальных клиентов когорты по периодам. Каждая ячейка содержит количество уникальных клиентов когорты, которые вернулись в любой период от начала когорты до текущего включительно."
                        view_key = "accumulation"
                        
                    elif view_type == "Динамика накопления возврата в %":
                        accumulation_percent_matrix = st.session_state.accumulation_percent_matrix
                        display_matrix = apply_matrix_color_gradient(accumulation_percent_matrix, hide_zeros=True, horizontal_dynamics=True, hide_before_diagonal=True)
                        
                        # Форматирование процентов
                        def format_percent_cell(val):
                            if pd.isna(val) or val == '':
                                return ''
                            try:
                                val_float = float(val)
                                if val_float == 0:
                                    return ''
                                return f"{val_float:.1f}%"
                            except (ValueError, TypeError):
                                if isinstance(val, str) and '%' in val:
                                    return val
                                return ''
                        
                        display_matrix = display_matrix.format(formatter=format_percent_cell)
                        description_text = "Показывает долю накопления уникальных клиентов когорты от общего количества клиентов в когорте. Значения выражены в процентах."
                        view_key = "accumulation_percent"
                        
                    elif view_type == "Приток возврата в %":
                        inflow_matrix = st.session_state.inflow_matrix
                        display_matrix = apply_matrix_color_gradient(inflow_matrix, hide_zeros=True, horizontal_dynamics=True, hide_before_diagonal=True)
                        
                        # Форматирование процентов для притока
                        def format_inflow_percent_cell(val):
                            if pd.isna(val) or val == '':
                                return ''
                            try:
                                val_float = float(val)
                                if val_float == 0:
                                    return ''
                                return f"{val_float:.1f}%"
                            except (ValueError, TypeError):
                                if isinstance(val, str) and '%' in val:
                                    return val
                                return ''
                        
                        # Добавляем 0.0% на диагонали
                        for row_name in display_matrix.data.index:
                            if row_name in display_matrix.data.columns:
                                display_matrix.data.loc[row_name, row_name] = '0.0%'
                        
                        format_dict_inflow = {col: format_inflow_percent_cell for col in display_matrix.data.columns}
                        display_matrix = display_matrix.format(format_dict_inflow)
                        description_text = "Показывает прирост уникальных клиентов когорты между периодами. Диагональ = 0%, первый период после диагонали = процент возврата, остальные = разница между накопительными процентами соседних периодов."
                        view_key = "inflow"
                    
                    elif view_type == "Отток клиентов из категории":
                        # Используем сохраненную таблицу оттока
                        if st.session_state.get('churn_table') is not None:
                            churn_table = st.session_state.churn_table
                            
                            # Форматируем таблицу для отображения
                            churn_display = churn_table.copy()
                            churn_display['Накопительный % возврата'] = churn_display['Накопительный % возврата'].apply(lambda x: x if x == '-' else f"{float(x):.1f}%")
                            churn_display['Отток %'] = churn_display['Отток %'].apply(lambda x: x if x == '-' else f"{float(x):.1f}%")
                            
                            # Убеждаемся, что когорта - первый столбец
                            column_order = ['Когорта', 'Кол-во клиентов когорты', 'Накопительное кол-во возврата', 
                                          'Накопительный % возврата', 'Отток кол-во', 'Отток %']
                            churn_display = churn_display[column_order]
                            
                            # Применяем стили для центрирования значений во всех столбцах
                            def center_format(val):
                                return 'text-align: center'
                            
                            # Создаем стилизованную таблицу с центрированием
                            styled_churn = churn_display.style.applymap(center_format)
                            
                            # Используем styled_churn как display_matrix для единообразия
                            display_matrix = styled_churn
                            description_text = "Показывает клиентов, которые не вернулись в категорию ни разу после периода когорты."
                            view_key = "churn"
                        else:
                            st.error("Таблица оттока не загружена. Пожалуйста, загрузите данные заново.")
                            display_matrix = None
                            description_text = ""
                            view_key = ""
                    
                    # Отображение описания с красивым оформлением
                    if description_text:
                        st.markdown(f'<div class="description-block">{description_text}</div>', unsafe_allow_html=True)
                    
                    # Создаем колонки для таблицы и кодов клиентов
                    col_table, col_clients = st.columns([4, 1])
                    
                    with col_table:
                        # Отображение таблицы (широкая) с поддержкой полноэкранного режима
                        if display_matrix is not None:
                            # Для таблицы оттока скрываем индекс
                            if view_key == "churn":
                                st.dataframe(
                                    display_matrix,
                                    use_container_width=True,
                                    hide_index=True
                                )
                                # Добавляем CSS для центрирования значений в таблице оттока
                                st.markdown("""
                                <style>
                                div[data-testid="stDataFrame"] table td {
                                    text-align: center !important;
                                }
                                div[data-testid="stDataFrame"] table th {
                                    text-align: center !important;
                                }
                                </style>
                                """, unsafe_allow_html=True)
                            else:
                                st.dataframe(
                                    display_matrix,
                                    use_container_width=True
                                )
                        else:
                            st.info("Выберите тип отображения для просмотра данных.")
                    
                    with col_clients:
                        # Компактный блок кодов клиентов
                        st.markdown('<div style="background: white; padding: 10px; border-radius: 8px; margin-bottom: 10px; border: 2px solid #ccc; box-shadow: 0 2px 4px rgba(0,0,0,0.1);"><h4 style="color: #333; margin: 0;">👥 Коды клиентов</h4></div>', unsafe_allow_html=True)
                        
                        # Коды клиентов в зависимости от выбранного типа
                        if view_key == "cohort":
                            selected_cohort = st.selectbox(
                                "Когорта:",
                                options=sorted_periods,
                                index=0,
                                help="Выберите период, когда клиенты впервые появились",
                                key="cohort_select_unified_1"
                            )
                            
                            selected_period = st.selectbox(
                                "Период:",
                                options=sorted_periods,
                                index=min(1, len(sorted_periods) - 1) if len(sorted_periods) > 1 else 0,
                                help="Выберите период, для которого нужно показать клиентов",
                                key="period_select_unified_1"
                            )
                            
                            if selected_cohort and selected_period:
                                period_clients_cache = st.session_state.get('period_clients_cache', None)
                                client_cohorts_cache = st.session_state.get('client_cohorts_cache', None)
                                common_clients = get_cohort_clients(df, year_month_col, client_col, selected_cohort, selected_period, period_clients_cache, client_cohorts_cache)
                                
                                if common_clients:
                                    st.write(f"**Найдено: {len(common_clients)}**")
                                    clients_csv = "\n".join([str(client) for client in common_clients])
                                    create_copy_button(
                                        clients_csv,
                                        f"📋 Копировать ({len(common_clients)})",
                                        "copy_clients_unified_1"
                                    )
                                else:
                                    st.info(f"❌ Нет данных")
                        
                        elif view_key == "accumulation":
                            selected_cohort = st.selectbox(
                                "Когорта:",
                                options=sorted_periods,
                                index=0,
                                help="Выберите период когорты",
                                key="cohort_select_unified_2"
                            )
                            
                            selected_period = st.selectbox(
                                "Период:",
                                options=sorted_periods,
                                index=min(1, len(sorted_periods) - 1) if len(sorted_periods) > 1 else 0,
                                help="Выберите период, до которого показывать накопленных клиентов",
                                key="period_select_unified_2"
                            )
                            
                            if selected_cohort and selected_period:
                                period_clients_cache = st.session_state.get('period_clients_cache', None)
                                client_cohorts_cache = st.session_state.get('client_cohorts_cache', None)
                                accumulation_clients = get_accumulation_clients(df, year_month_col, client_col, sorted_periods, selected_cohort, selected_period, period_clients_cache=period_clients_cache, client_cohorts_cache=client_cohorts_cache)
                                
                                if accumulation_clients:
                                    st.write(f"**Найдено: {len(accumulation_clients)}**")
                                    clients_csv = "\n".join([str(client) for client in accumulation_clients])
                                    create_copy_button(
                                        clients_csv,
                                        f"📋 Копировать ({len(accumulation_clients)})",
                                        "copy_clients_unified_2"
                                    )
                                else:
                                    st.info(f"❌ Нет данных")
                        
                        elif view_key == "accumulation_percent":
                            selected_cohort = st.selectbox(
                                "Когорта:",
                                options=sorted_periods,
                                index=0,
                                help="Выберите период когорты",
                                key="cohort_select_unified_3"
                            )
                            
                            selected_period = st.selectbox(
                                "Период:",
                                options=sorted_periods,
                                index=min(1, len(sorted_periods) - 1) if len(sorted_periods) > 1 else 0,
                                help="Выберите период, до которого показывать накопленных клиентов",
                                key="period_select_unified_3"
                            )
                            
                            if selected_cohort and selected_period:
                                period_clients_cache = st.session_state.get('period_clients_cache', None)
                                client_cohorts_cache = st.session_state.get('client_cohorts_cache', None)
                                accumulation_clients = get_accumulation_clients(df, year_month_col, client_col, sorted_periods, selected_cohort, selected_period, period_clients_cache=period_clients_cache, client_cohorts_cache=client_cohorts_cache)
                                
                                if accumulation_clients:
                                    st.write(f"**Найдено: {len(accumulation_clients)}**")
                                    clients_csv = "\n".join([str(client) for client in accumulation_clients])
                                    create_copy_button(
                                        clients_csv,
                                        f"📋 Копировать ({len(accumulation_clients)})",
                                        "copy_clients_unified_3"
                                    )
                                else:
                                    st.info(f"❌ Нет данных")
                        
                        elif view_key == "inflow":
                            selected_cohort = st.selectbox(
                                "Когорта:",
                                options=sorted_periods,
                                index=0,
                                help="Выберите период когорты",
                                key="cohort_select_unified_4"
                            )
                            
                            selected_period = st.selectbox(
                                "Период:",
                                options=sorted_periods,
                                index=min(1, len(sorted_periods) - 1) if len(sorted_periods) > 1 else 0,
                                help="Выберите период, для которого показать новых вернувшихся клиентов",
                                key="period_select_unified_4"
                            )
                            
                            if selected_cohort and selected_period:
                                period_clients_cache = st.session_state.get('period_clients_cache', None)
                                client_cohorts_cache = st.session_state.get('client_cohorts_cache', None)
                                inflow_clients = get_inflow_clients(df, year_month_col, client_col, sorted_periods, selected_cohort, selected_period, period_clients_cache, client_cohorts_cache)
                                
                                if inflow_clients:
                                    st.write(f"**Найдено: {len(inflow_clients)}**")
                                    clients_csv = "\n".join([str(client) for client in inflow_clients])
                                    create_copy_button(
                                        clients_csv,
                                        f"📋 Копировать ({len(inflow_clients)})",
                                        "copy_clients_unified_4"
                                    )
                                else:
                                    st.info(f"❌ Нет данных")
                        
                        elif view_key == "churn":
                            # Для оттока только выбор когорты, без периода
                            selected_cohort = st.selectbox(
                                "Когорта:",
                                options=sorted_periods,
                                index=0,
                                help="Выберите когорту для скачивания списка клиентов оттока из категории",
                                key="cohort_select_unified_5"
                            )
                            
                            if selected_cohort:
                                period_clients_cache = st.session_state.get('period_clients_cache', None)
                                client_cohorts_cache = st.session_state.get('client_cohorts_cache', None)
                                churn_clients = get_churn_clients(df, year_month_col, client_col, sorted_periods, selected_cohort, period_clients_cache, client_cohorts_cache)
                                
                                if churn_clients:
                                    st.write(f"**Найдено: {len(churn_clients)}**")
                                    clients_csv = "\n".join([str(client) for client in churn_clients])
                                    create_copy_button(
                                        clients_csv,
                                        f"📋 Копировать ({len(churn_clients)})",
                                        "copy_clients_unified_5"
                                    )
                                else:
                                    st.info(f"❌ Нет данных")
                                
                                # Кнопка для скачивания всех когорт (всегда видна)
                                all_churn_clients = set()
                                client_cohorts_cache = st.session_state.get('client_cohorts_cache', None)
                                for cohort in sorted_periods:
                                    cohort_churn = get_churn_clients(df, year_month_col, client_col, sorted_periods, cohort, period_clients_cache, client_cohorts_cache)
                                    all_churn_clients.update(cohort_churn)
                                
                                if all_churn_clients:
                                    all_clients_csv = "\n".join([str(client) for client in sorted(all_churn_clients)])
                                    create_copy_button(
                                        all_clients_csv,
                                        f"📋 Копировать коды клиентов оттока всех когорт ({len(all_churn_clients)})",
                                        "copy_all_churn_clients"
                                    )
                    
                    # Шестой блок - Присутствие клиентов оттока в других категориях
                    st.markdown("---")
                    
                    # Блок с заголовками, инструкциями и изображением в одной строке
                    col_churn_title_left, col_churn_title_right = st.columns([1, 1])
                    
                    with col_churn_title_left:
                        st.subheader("🔍 Присутствие клиентов оттока в других категориях и отток из сети")
                        # Текст инструкций прямо под заголовком
                        st.markdown("""
                        1) Скопируйте коды клиентов оттока всех когорт в блоке "Отток клиентов из категории"
                        
                        2) Примените коды клиентов. Отберите анализируемый период и разрез ( год-месяц / год - неделя) идентичный построенному когортному анализу.
                        
                        3) Отберите анализируемые категории в одном из разрезов Группа1/2/3/4.
                        
                        4) Зайдите на лист "Конструктор" и выведите отчёт по шаблону справа.
                        
                        5) Скачайте документ в Qlik и загрузите в ячейку справа.
                        """)
                    
                    with col_churn_title_right:
                        st.subheader("📋 Шаблон загрузки данных из Qlik")
                        # Пытаемся найти скриншот шаблона для категорий (изображение сразу под заголовком)
                        churn_categories_image_paths = [
                            'qlik_template_categories.png',
                            'qlik_template_categories.jpg',
                            'qlik_template_categories.jpeg',
                            'шаблон_qlik_категории.png',
                            'шаблон_qlik_категории.jpg',
                            'шаблон_qlik_категории.jpeg',
                            'churn_categories_template.png',
                            'churn_categories_template.jpg',
                            'churn_categories_template.jpeg'
                        ]
                        image_found = False
                        for img_path in churn_categories_image_paths:
                            if os.path.exists(img_path):
                                st.image(img_path, use_container_width=True)
                                image_found = True
                                break
                        if not image_found:
                            st.info("📸 Поместите скриншот шаблона загрузки данных из Qlik в папку проекта с одним из имён: qlik_template_categories.png, шаблон_qlik_категории.png или churn_categories_template.png")
                        
                        # Загрузчик Excel файла прямо под изображением
                        uploaded_file_categories = st.file_uploader(
                            "Выберите Excel файл с данными о присутствии клиентов оттока в других категориях",
                            type=['xlsx', 'xls'],
                            help="Загрузите файл, скачанный из Qlik согласно шаблону выше",
                            key="upload_categories_file"
                        )
                    
                    # Обработка загруженного файла
                    if uploaded_file_categories is not None:
                        try:
                            # Загрузка Excel файла
                            if uploaded_file_categories.name.endswith('.xlsx'):
                                df_categories = pd.read_excel(uploaded_file_categories, engine='openpyxl')
                            else:
                                df_categories = pd.read_excel(uploaded_file_categories, engine='xlrd')
                            
                            # Определяем столбцы
                            group_col = None
                            year_month_col = None
                            month_col = None
                            clients_col = None
                            client_code_col = None
                            
                            # Ищем столбец Группа (может быть Группа1, Группа2, Группа3 и т.д.)
                            for col in df_categories.columns:
                                col_lower = str(col).lower().strip()
                                if 'группа' in col_lower:
                                    group_col = col
                                    break
                            
                            # Ищем столбец периода (Год-месяц или Год-неделя)
                            for col in df_categories.columns:
                                col_lower = str(col).lower().strip()
                                if 'год' in col_lower and ('месяц' in col_lower or 'неделя' in col_lower):
                                    year_month_col = col
                                    break
                            
                            # Ищем столбец месяц
                            for col in df_categories.columns:
                                col_lower = str(col).lower().strip()
                                if col_lower == 'месяц' or (col_lower.startswith('месяц') and len(col_lower.split()) == 1):
                                    month_col = col
                                    break
                            
                            # Ищем столбец Клиентов
                            for col in df_categories.columns:
                                col_lower = str(col).lower().strip()
                                if 'клиент' in col_lower and ('ов' in col_lower or 'ов' in col_lower):
                                    clients_col = col
                                    break
                            
                            # Ищем столбец Код клиента
                            for col in df_categories.columns:
                                col_lower = str(col).lower().strip()
                                if 'код' in col_lower and 'клиент' in col_lower:
                                    client_code_col = col
                                    break
                            
                            # Проверяем наличие всех необходимых столбцов
                            if group_col is None:
                                st.error("❌ Не найден столбец с категориями (Группа1, Группа2, Группа3 и т.д.). Убедитесь, что в файле есть столбец с названием, содержащим 'Группа'.")
                            elif client_code_col is None:
                                st.error("❌ Не найден столбец 'Код клиента'. Убедитесь, что в файле есть столбец с названием, содержащим 'Код' и 'клиент'.")
                            elif year_month_col is None:
                                st.warning("⚠️ Не найден столбец периода ('Год-месяц' или 'Год-неделя'). Данные будут обработаны без фильтрации по периоду.")
                            else:
                                # Получаем уникальные категории
                                categories = df_categories[group_col].dropna().unique()
                                categories = sorted([str(cat) for cat in categories if str(cat).strip() != ''])
                                
                                # Сохраняем данные о категориях в session_state для использования в Excel отчёте
                                st.session_state.df_categories = df_categories
                                st.session_state.categories_list = categories
                                st.session_state.group_col_name = group_col
                                st.session_state.year_month_col_name = year_month_col
                                st.session_state.client_code_col_name = client_code_col
                                
                                # Получаем клиентов оттока для каждой когорты
                                period_clients_cache = st.session_state.get('period_clients_cache', None)
                                
                                # Рассчитываем метрики для всех когорт для сводной таблицы
                                total_present_after_cohort_by_cohort = {}
                                total_present_after_cohort_percent_by_cohort = {}
                                network_churn_by_cohort = {}
                                network_churn_percent_by_cohort = {}
                                
                                # Собираем всех клиентов из категорий (для всех периодов)
                                all_category_clients_all_periods = set()
                                if year_month_col is not None:
                                    for category in categories:
                                        category_data = df_categories[df_categories[group_col] == category]
                                        category_clients = set(category_data[client_code_col].dropna().astype(str).unique())
                                        all_category_clients_all_periods.update(category_clients)
                                else:
                                    for category in categories:
                                        category_data = df_categories[df_categories[group_col] == category]
                                        category_clients = set(category_data[client_code_col].dropna().astype(str).unique())
                                        all_category_clients_all_periods.update(category_clients)
                                
                                # Для каждой когорты рассчитываем метрики
                                churn_table = st.session_state.churn_table
                                client_cohorts_cache = st.session_state.get('client_cohorts_cache', None)
                                # Собираем всех клиентов оттока из сети для всех когорт
                                all_network_churn_clients = set()
                                for cohort_period in sorted_periods:
                                    # Получаем клиентов оттока для этой когорты
                                    churn_clients_set_cohort = set(get_churn_clients(df, year_month_col, client_col, sorted_periods, cohort_period, period_clients_cache, client_cohorts_cache))
                                    churn_clients_set_cohort = {str(client) for client in churn_clients_set_cohort}
                                    
                                    # Получаем отток из категории для этой когорты
                                    cohort_row = churn_table[churn_table['Когорта'] == cohort_period]
                                    churn_count_cohort = _churn_int(cohort_row.iloc[0]['Отток кол-во']) if not cohort_row.empty else 0
                                    cohort_size_cohort = int(cohort_row.iloc[0]['Кол-во клиентов когорты']) if not cohort_row.empty else 0
                                    
                                    # Определяем периоды начиная с этой когорты
                                    cohort_index_cohort = sorted_periods.index(cohort_period) if cohort_period in sorted_periods else 0
                                    periods_from_cohort_cohort = sorted_periods[cohort_index_cohort:]
                                    # Периоды ПОСЛЕ когорты (исключая период когорты)
                                    periods_after_cohort_cohort = periods_from_cohort_cohort[1:] if len(periods_from_cohort_cohort) > 1 else []
                                    
                                    # Клиенты оттока, присутствующие в других категориях ПОСЛЕ месяца когорты
                                    all_category_clients_after_cohort = set()
                                    if year_month_col is not None and len(periods_after_cohort_cohort) > 0:
                                        for category in categories:
                                            category_data = df_categories[df_categories[group_col] == category]
                                            category_data_filtered = category_data[category_data[year_month_col].isin(periods_after_cohort_cohort)]
                                            category_clients = set(category_data_filtered[client_code_col].dropna().astype(str).unique())
                                            all_category_clients_after_cohort.update(category_clients)
                                    elif year_month_col is None:
                                        all_category_clients_after_cohort = all_category_clients_all_periods
                                    
                                    present_in_categories_after_cohort = churn_clients_set_cohort & all_category_clients_after_cohort
                                    total_present_after_cohort_by_cohort[cohort_period] = len(present_in_categories_after_cohort)
                                    
                                    # % присутствия после месяца когорты
                                    present_after_cohort_percent = (len(present_in_categories_after_cohort) / cohort_size_cohort * 100) if cohort_size_cohort > 0 else 0
                                    total_present_after_cohort_percent_by_cohort[cohort_period] = present_after_cohort_percent
                                    
                                    # Отток из сети = Отток из категории - Клиентов когорты присутствуют в других категориях после месяца когорты
                                    network_churn_cohort = churn_count_cohort - len(present_in_categories_after_cohort)
                                    network_churn_by_cohort[cohort_period] = max(0, network_churn_cohort)  # Не может быть отрицательным
                                    
                                    # % оттока из сети
                                    network_churn_percent_cohort = (network_churn_by_cohort[cohort_period] / cohort_size_cohort * 100) if cohort_size_cohort > 0 else 0
                                    network_churn_percent_by_cohort[cohort_period] = network_churn_percent_cohort
                                    
                                    # Собираем клиентов оттока из сети для этой когорты
                                    network_churn_clients_cohort = churn_clients_set_cohort - all_category_clients_after_cohort
                                    all_network_churn_clients.update(network_churn_clients_cohort)
                                
                                # Ключи метрик с учётом типа периода (недели/месяцы)
                                _pa = st.session_state.get('period_after_label', 'месяца')
                                _key_итого = f"Итого присутствуют в других категориях после {_pa} когорты"
                                _key_доля = f"Доля присутствуют в других категориях после {_pa} когорты"
                                summary_table_excel = pd.DataFrame({
                                    'Отток из сети': network_churn_by_cohort,
                                    'Доля оттока из сети от когорты': network_churn_percent_by_cohort,
                                    _key_итого: total_present_after_cohort_by_cohort,
                                    _key_доля: total_present_after_cohort_percent_by_cohort
                                })
                                summary_table_excel = summary_table_excel.T
                                
                                # Сохраняем данные для Excel отчёта и сводной таблицы
                                st.session_state.category_summary_table = summary_table_excel
                                st.session_state.category_cohort_table = None
                                
                                # Устанавливаем флаг успешной загрузки и обработки второго файла
                                st.session_state.categories_file_uploaded = True
                                
                                # Обновляем Excel отчёт после сохранения всех данных
                                if 'excel_report_cache_key' in st.session_state:
                                    del st.session_state.excel_report_cache_key
                                
                                # Перегенерируем Excel отчёт после сохранения данных о категориях
                                # Используем st.rerun() для обновления, но только если данные изменились
                                # Вместо этого просто перегенерируем отчет
                                try:
                                    # Небольшая задержка для гарантии сохранения данных
                                    st.session_state.excel_report_data = create_full_report_excel()
                                except Exception as e:
                                    st.warning(f"Не удалось обновить Excel отчёт: {str(e)}")
                                
                                # Новый интерфейс: слева выбор когорты, справа таблица
                                st.markdown("### 📊 Присутствие клиентов оттока когорты в других категориях товаров")
                                
                                col_cohort_select, col_table = st.columns([1, 4])
                                
                                with col_cohort_select:
                                    selected_cohort = st.selectbox(
                                        "Выберите когорту:",
                                        options=sorted_periods,
                                        index=0,
                                        help="Выберите когорту для анализа присутствия её клиентов оттока в других категориях",
                                        key="category_cohort_select"
                                    )
                                    
                                    # Определяем периоды начиная с выбранной когорты
                                    cohort_index = sorted_periods.index(selected_cohort) if selected_cohort in sorted_periods else 0
                                    periods_from_cohort = sorted_periods[cohort_index:]
                                    # Периоды ПОСЛЕ когорты (исключая период когорты) - начинаем расчет с этого периода
                                    periods_after_cohort = periods_from_cohort[1:] if len(periods_from_cohort) > 1 else []
                                    
                                    # Получаем клиентов оттока для выбранной когорты
                                    client_cohorts_cache = st.session_state.get('client_cohorts_cache', None)
                                    churn_clients_set = set(get_churn_clients(df, year_month_col, client_col, sorted_periods, selected_cohort, period_clients_cache, client_cohorts_cache))
                                    churn_clients_set = {str(client) for client in churn_clients_set}
                                    
                                    # Получаем размер когорты и отток из churn_table
                                    churn_table = st.session_state.churn_table
                                    cohort_row = churn_table[churn_table['Когорта'] == selected_cohort]
                                    cohort_size = int(cohort_row.iloc[0]['Кол-во клиентов когорты']) if not cohort_row.empty else 0
                                    churn_count = _churn_int(cohort_row.iloc[0]['Отток кол-во']) if not cohort_row.empty else 0
                                    
                                    # Клиенты оттока, присутствующие в других категориях ПОСЛЕ месяца когорты
                                    all_category_clients_after_cohort = set()
                                    if year_month_col is not None and len(periods_after_cohort) > 0:
                                        # Используем только данные из периодов ПОСЛЕ выбранной когорты
                                        for category in categories:
                                            category_data = df_categories[df_categories[group_col] == category]
                                            # Фильтруем только периоды ПОСЛЕ выбранной когорты
                                            category_data_filtered = category_data[category_data[year_month_col].isin(periods_after_cohort)]
                                            category_clients = set(category_data_filtered[client_code_col].dropna().astype(str).unique())
                                            all_category_clients_after_cohort.update(category_clients)
                                    elif year_month_col is None:
                                        # Если нет столбца "Год-месяц", собираем всех клиентов из всех категорий
                                        for category in categories:
                                            category_data = df_categories[df_categories[group_col] == category]
                                            category_clients = set(category_data[client_code_col].dropna().astype(str).unique())
                                            all_category_clients_after_cohort.update(category_clients)
                                    
                                    present_in_categories_after_cohort = churn_clients_set & all_category_clients_after_cohort
                                    present_count_after_cohort = len(present_in_categories_after_cohort)
                                    present_percent_after_cohort = (present_count_after_cohort / cohort_size * 100) if cohort_size > 0 else 0
                                    
                                    # Отток из сети = Отток из категории - Клиентов когорты присутствуют в других категориях после месяца когорты
                                    network_churn = churn_count - present_count_after_cohort
                                    network_churn = max(0, network_churn)  # Не может быть отрицательным
                                    network_churn_percent = (network_churn / cohort_size * 100) if cohort_size > 0 else 0
                                    
                                    # Вычисляем клиентов оттока из сети
                                    # Это клиенты оттока, которые НЕ присутствуют в других категориях после месяца когорты
                                    network_churn_clients = churn_clients_set - all_category_clients_after_cohort
                                    network_churn_clients_list = sorted(list(network_churn_clients))
                                    
                                    _pa_label = st.session_state.get('period_after_label', 'месяца')
                                    metrics_html = f"""
                                    <div style="line-height: 2;">
                                    <p style="color: #333; font-size: 1rem; margin: 8px 0;">
                                        <strong style="color: #1f77b4;">Клиентов когорты присутствуют в других категориях после {_pa_label} когорты:</strong> 
                                        <span style="color: #2c3e50; font-weight: 600;">{present_count_after_cohort} ({present_percent_after_cohort:.1f}%)</span>
                                    </p>
                                    <p style="color: #333; font-size: 1rem; margin: 8px 0;">
                                        <strong style="color: #1f77b4;">Отток из сети:</strong> 
                                        <span style="color: #e74c3c; font-weight: 600;">{network_churn} ({network_churn_percent:.1f}%)</span>
                                    </p>
                                    </div>
                                    """
                                    st.markdown(metrics_html, unsafe_allow_html=True)
                                    
                                    # Кнопка копирования кодов клиентов оттока из сети для выбранной когорты
                                    if network_churn_clients_list:
                                        network_churn_clients_csv = "\n".join([str(client) for client in network_churn_clients_list])
                                        create_copy_button(
                                            network_churn_clients_csv,
                                            f"📋 Копировать коды клиентов оттока из сети ({len(network_churn_clients_list)})",
                                            f"copy_network_churn_{selected_cohort}"
                                        )
                                    else:
                                        st.info("ℹ️ Отток из сети равен 0 или все клиенты оттока присутствуют в других категориях")
                                
                                with col_table:
                                    # Определяем периоды ПОСЛЕ выбранной когорты (для использования в таблице)
                                    cohort_index_table = sorted_periods.index(selected_cohort) if selected_cohort in sorted_periods else 0
                                    periods_from_cohort_table = sorted_periods[cohort_index_table:]
                                    # Периоды ПОСЛЕ когорты (исключая период когорты) - начинаем расчет с этого периода
                                    periods_after_cohort_table = periods_from_cohort_table[1:] if len(periods_from_cohort_table) > 1 else []
                                    
                                    # Создаем таблицу: категории по строкам, периоды по столбцам (только ПОСЛЕ выбранной когорты)
                                    category_period_table = pd.DataFrame(index=categories, columns=periods_after_cohort_table)
                                    
                                    # Словарь для хранения уникальных клиентов по периодам (для итоговой строки)
                                    period_unique_clients = {period: set() for period in periods_after_cohort_table}
                                    
                                    # Словарь для хранения уникальных клиентов по категориям (для итогового столбца)
                                    category_unique_clients = {category: set() for category in categories}
                                    
                                    # Если есть столбец "Год-месяц", используем его для фильтрации по периодам
                                    if year_month_col is not None:
                                        # Для каждого периода ПОСЛЕ выбранной когорты проверяем присутствие клиентов оттока в категориях
                                        for period in periods_after_cohort_table:
                                            # Фильтруем данные по периоду
                                            period_data = df_categories[df_categories[year_month_col] == period]
                                            
                                            # Для каждой категории считаем количество клиентов оттока, присутствующих в этом периоде
                                            for category in categories:
                                                # Данные категории в этом периоде
                                                category_period_data = period_data[
                                                    (period_data[group_col] == category) & 
                                                    (period_data[client_code_col].notna())
                                                ]
                                                
                                                # Коды клиентов этой категории в этом периоде
                                                category_period_clients = set(
                                                    category_period_data[client_code_col].dropna().astype(str).unique()
                                                )
                                                
                                                # Находим пересечение: клиенты оттока выбранной когорты, которые есть в этой категории в этом периоде
                                                intersection = churn_clients_set & category_period_clients
                                                category_period_table.loc[category, period] = len(intersection)
                                                
                                                # Добавляем в множества для итогов
                                                period_unique_clients[period].update(intersection)
                                                category_unique_clients[category].update(intersection)
                                    else:
                                        # Если нет столбца "Год-месяц", используем все данные без фильтрации по периоду
                                        # Создаем словарь: категория -> множество кодов клиентов
                                        category_clients_dict = {}
                                        for category in categories:
                                            category_data = df_categories[df_categories[group_col] == category]
                                            client_codes = set(category_data[client_code_col].dropna().astype(str).unique())
                                            category_clients_dict[category] = client_codes
                                        
                                        # Для каждого периода ПОСЛЕ выбранной когорты используем одинаковые данные
                                        for period in periods_after_cohort_table:
                                            for category in categories:
                                                category_clients_set = category_clients_dict.get(category, set())
                                                intersection = churn_clients_set & category_clients_set
                                                category_period_table.loc[category, period] = len(intersection)
                                                
                                                # Добавляем в множества для итогов
                                                period_unique_clients[period].update(intersection)
                                                category_unique_clients[category].update(intersection)
                                    
                                    # Заполняем NaN нулями
                                    category_period_table = category_period_table.fillna(0).astype(int)
                                    
                                    # Создаем итоговую строку по периодам (уникальные клиенты по всем категориям)
                                    totals_row = pd.Series(
                                        {period: len(period_unique_clients[period]) for period in periods_after_cohort_table},
                                        name='Итого клиентов'
                                    )
                                    
                                    # Создаем итоговый столбец по категориям (уникальные клиенты за весь период)
                                    totals_col = pd.Series(
                                        {category: len(category_unique_clients[category]) for category in categories},
                                        name='Итого'
                                    )
                                    
                                    # Добавляем итоговую строку в таблицу
                                    category_period_table_with_totals = category_period_table.copy()
                                    category_period_table_with_totals.loc['Итого клиентов'] = totals_row
                                    
                                    # Добавляем итоговый столбец
                                    category_period_table_with_totals['Итого'] = totals_col
                                    category_period_table_with_totals.loc['Итого клиентов', 'Итого'] = len(present_in_categories_after_cohort)
                                    
                                    # Переупорядочиваем строки: итоговая строка наверх
                                    new_index = ['Итого клиентов'] + [cat for cat in categories]
                                    category_period_table_with_totals = category_period_table_with_totals.reindex(new_index)
                                    
                                    # Переупорядочиваем столбцы: итоговый столбец слева (после индекса, перед периодами)
                                    new_columns = ['Итого'] + list(periods_after_cohort_table)
                                    category_period_table_with_totals = category_period_table_with_totals[new_columns]
                                    
                                    # Отображаем основную таблицу с итогами
                                    st.dataframe(
                                        category_period_table_with_totals,
                                        use_container_width=True
                                    )
                                    
                                    # Добавляем стили для центрирования, выделения итоговых значений жирным, пастельным цветом и закрепления
                                    st.markdown("""
                                    <style>
                                    div[data-testid="stDataFrame"] table td {
                                        text-align: center !important;
                                    }
                                    div[data-testid="stDataFrame"] table th {
                                        text-align: center !important;
                                    }
                                    /* Закрепляем первую строку (итоговая строка "Итого клиентов") сверху */
                                    div[data-testid="stDataFrame"] table tbody tr:first-child td,
                                    div[data-testid="stDataFrame"] table tbody tr:first-child th {
                                        font-weight: bold !important;
                                        background-color: #E3F2FD !important;
                                        position: sticky !important;
                                        top: 0 !important;
                                        z-index: 10 !important;
                                    }
                                    /* Закрепляем первый столбец данных (итоговый столбец "Итого") слева */
                                    div[data-testid="stDataFrame"] table tbody tr td:nth-child(2),
                                    div[data-testid="stDataFrame"] table thead tr th:nth-child(2) {
                                        font-weight: bold !important;
                                        background-color: #E3F2FD !important;
                                        position: sticky !important;
                                        left: 0 !important;
                                        z-index: 5 !important;
                                    }
                                    /* Закрепляем ячейку пересечения итоговых строки и столбца (и сверху, и слева) */
                                    div[data-testid="stDataFrame"] table tbody tr:first-child td:nth-child(2) {
                                        background-color: #BBDEFB !important;
                                        font-weight: bold !important;
                                        position: sticky !important;
                                        top: 0 !important;
                                        left: 0 !important;
                                        z-index: 15 !important;
                                    }
                                    /* Закрепляем заголовок итогового столбца */
                                    div[data-testid="stDataFrame"] table thead tr th:nth-child(2) {
                                        position: sticky !important;
                                        left: 0 !important;
                                        z-index: 6 !important;
                                    }
                                    </style>
                                    <script>
                                    // Дополнительный скрипт для гарантированного выделения жирным, цветом и закрепления
                                    setTimeout(function() {
                                        const tables = document.querySelectorAll('div[data-testid="stDataFrame"] table');
                                        tables.forEach(table => {
                                            // Первая строка (итоговая) - закрепляем сверху
                                            const firstRow = table.querySelector('tbody tr:first-child');
                                            if (firstRow) {
                                                firstRow.querySelectorAll('td, th').forEach(cell => {
                                                    cell.style.fontWeight = 'bold';
                                                    cell.style.position = 'sticky';
                                                    cell.style.top = '0';
                                                    cell.style.zIndex = '10';
                                                    if (!cell.style.backgroundColor || cell.style.backgroundColor === '') {
                                                        cell.style.backgroundColor = '#E3F2FD';
                                                    }
                                                });
                                            }
                                            // Первый столбец данных (итоговый) - закрепляем слева
                                            table.querySelectorAll('tbody tr').forEach(row => {
                                                const firstDataCell = row.querySelector('td:nth-child(2)');
                                                if (firstDataCell) {
                                                    firstDataCell.style.fontWeight = 'bold';
                                                    firstDataCell.style.position = 'sticky';
                                                    firstDataCell.style.left = '0';
                                                    firstDataCell.style.zIndex = '5';
                                                    if (!firstDataCell.style.backgroundColor || firstDataCell.style.backgroundColor === '') {
                                                        firstDataCell.style.backgroundColor = '#E3F2FD';
                                                    }
                                                }
                                            });
                                            const firstHeader = table.querySelector('thead th:nth-child(2)');
                                            if (firstHeader) {
                                                firstHeader.style.fontWeight = 'bold';
                                                firstHeader.style.backgroundColor = '#E3F2FD';
                                                firstHeader.style.position = 'sticky';
                                                firstHeader.style.left = '0';
                                                firstHeader.style.zIndex = '6';
                                            }
                                            // Ячейка пересечения - закрепляем и сверху, и слева
                                            const intersectionCell = table.querySelector('tbody tr:first-child td:nth-child(2)');
                                            if (intersectionCell) {
                                                intersectionCell.style.backgroundColor = '#BBDEFB';
                                                intersectionCell.style.position = 'sticky';
                                                intersectionCell.style.top = '0';
                                                intersectionCell.style.left = '0';
                                                intersectionCell.style.zIndex = '15';
                                            }
                                        });
                                    }, 100);
                                    </script>
                                    """, unsafe_allow_html=True)
                                
                        except Exception as e:
                            st.error(f"❌ Ошибка при обработке файла: {str(e)}")
                            st.exception(e)
                            # Сбрасываем флаг при ошибке обработки
                            st.session_state.categories_file_uploaded = False
                    else:
                        # Если файл не загружен, сбрасываем флаг и очищаем данные
                        if st.session_state.get('categories_file_uploaded', False):
                            st.session_state.categories_file_uploaded = False
                            # Очищаем данные категорий
                            if 'df_categories' in st.session_state:
                                del st.session_state.df_categories
                            if 'category_summary_table' in st.session_state:
                                del st.session_state.category_summary_table
                            if 'category_cohort_table' in st.session_state:
                                del st.session_state.category_cohort_table
                    
                    # Сводная таблица по всем когортам (после блока присутствия клиентов)
                    st.markdown("---")
                    st.subheader("📊 Сводная таблица по всем когортам")
                    st.caption("Чем ближе когорта к последнему периоду в выгрузке, тем менее сопоставимы метрики: накопленный возврат ещё не успевает сформироваться, а доля оттока завышена из‑за короткого горизонта наблюдения.")
                    if st.session_state.get('churn_table') is not None:
                        churn_table = st.session_state.churn_table
                        has_categories_file = (
                            st.session_state.get('upload_categories_file') is not None or
                            st.session_state.get('category_summary_table') is not None
                        )
                        summary_data = {}
                        
                        # 1. Кол-во клиентов в когорте
                        summary_data['Кол-во клиентов в когорте'] = {}
                        for _, row in churn_table.iterrows():
                            cohort = row['Когорта']
                            summary_data['Кол-во клиентов в когорте'][cohort] = int(row['Кол-во клиентов когорты'])
                        
                        # 2. Накопительное кол-во вернувшихся в категорию
                        summary_data['Накопительное кол-во вернувшихся в категорию'] = {}
                        for _, row in churn_table.iterrows():
                            cohort = row['Когорта']
                            summary_data['Накопительное кол-во вернувшихся в категорию'][cohort] = _churn_int(row['Накопительное кол-во возврата'])
                        
                        # 3. Накопительное кол-во вернувшихся в категорию %
                        summary_data['Накопительное кол-во вернувшихся в категорию %'] = {}
                        for _, row in churn_table.iterrows():
                            cohort = row['Когорта']
                            v_ret = row['Накопительный % возврата']
                            summary_data['Накопительное кол-во вернувшихся в категорию %'][cohort] = v_ret if v_ret == '-' else f"{float(v_ret):.1f}%"
                        
                        # 4–5. Отток из категории (из первого файла — всегда при наличии churn_table)
                        summary_data['Отток из категории когорты'] = {}
                        for _, row in churn_table.iterrows():
                            cohort = row['Когорта']
                            summary_data['Отток из категории когорты'][cohort] = _churn_int(row['Отток кол-во'])
                        summary_data['Отток из категории когорты %'] = {}
                        for _, row in churn_table.iterrows():
                            cohort = row['Когорта']
                            v = row['Отток %']
                            summary_data['Отток из категории когорты %'][cohort] = v if v == '-' else f"{float(v):.1f}%"
                        
                        if has_categories_file:
                            _pa_ui = st.session_state.get('period_after_label', 'месяца')
                            _key_итого_ui = f"Итого присутствуют в других категориях после {_pa_ui} когорты"
                            _key_доля_ui = f"Доля присутствуют в других категориях после {_pa_ui} когорты"
                            _key_кол_ui = f"Кол-во клиентов когорты в других категориях после {_pa_ui} когорты"
                            _key_кол_pct_ui = f"Кол-во клиентов когорты в других категориях после {_pa_ui} когорты %"
                            summary_data[_key_кол_ui] = {}
                            summary_data[_key_кол_pct_ui] = {}
                            summary_data['Отток из сети'] = {}
                            summary_data['Отток из сети %'] = {}
                            
                            for cohort in sorted_periods:
                                summary_data[_key_кол_ui][cohort] = 0
                                summary_data[_key_кол_pct_ui][cohort] = "0.0%"
                                summary_data['Отток из сети'][cohort] = 0
                                summary_data['Отток из сети %'][cohort] = "0.0%"
                            
                            if 'category_summary_table' in st.session_state and st.session_state.category_summary_table is not None:
                                category_summary = st.session_state.category_summary_table
                                if _key_итого_ui in category_summary.index:
                                    for cohort in sorted_periods:
                                        if cohort in category_summary.columns:
                                            value = category_summary.loc[_key_итого_ui, cohort]
                                            summary_data[_key_кол_ui][cohort] = int(value) if pd.notna(value) else 0
                                        else:
                                            summary_data[_key_кол_ui][cohort] = 0
                                if _key_доля_ui in category_summary.index:
                                    for cohort in sorted_periods:
                                        if cohort in category_summary.columns:
                                            value = category_summary.loc[_key_доля_ui, cohort]
                                            if pd.notna(value):
                                                summary_data[_key_кол_pct_ui][cohort] = f"{float(value):.1f}%"
                                            else:
                                                summary_data[_key_кол_pct_ui][cohort] = "0.0%"
                                        else:
                                            summary_data[_key_кол_pct_ui][cohort] = "0.0%"
                                else:
                                    for cohort in sorted_periods:
                                        cohort_size = summary_data['Кол-во клиентов в когорте'].get(cohort, 0)
                                        present_after_count = summary_data[_key_кол_ui].get(cohort, 0)
                                        if cohort_size > 0:
                                            percent = (present_after_count / cohort_size) * 100
                                            summary_data[_key_кол_pct_ui][cohort] = f"{percent:.1f}%"
                                        else:
                                            summary_data[_key_кол_pct_ui][cohort] = "0.0%"
                                
                                # 8. Отток из сети (обновляем значения)
                                if 'Отток из сети' in category_summary.index:
                                    for cohort in sorted_periods:
                                        if cohort in category_summary.columns:
                                            value = category_summary.loc['Отток из сети', cohort]
                                            summary_data['Отток из сети'][cohort] = int(value) if pd.notna(value) else 0
                                        else:
                                            summary_data['Отток из сети'][cohort] = 0
                                else:
                                    for cohort in sorted_periods:
                                        summary_data['Отток из сети'][cohort] = 0
                                
                                # 9. Отток из сети % (обновляем значения)
                                if 'Доля оттока из сети от когорты' in category_summary.index:
                                    for cohort in sorted_periods:
                                        if cohort in category_summary.columns:
                                            value = category_summary.loc['Доля оттока из сети от когорты', cohort]
                                            if pd.notna(value):
                                                summary_data['Отток из сети %'][cohort] = f"{value:.1f}%"
                                            else:
                                                summary_data['Отток из сети %'][cohort] = "0.0%"
                                        else:
                                            summary_data['Отток из сети %'][cohort] = "0.0%"
                                else:
                                    # Если нет строки "Доля оттока из сети от когорты", вычисляем процент вручную
                                    for cohort in sorted_periods:
                                        cohort_size = summary_data['Кол-во клиентов в когорте'].get(cohort, 0)
                                        network_churn = summary_data['Отток из сети'].get(cohort, 0)
                                        if cohort_size > 0:
                                            percent = (network_churn / cohort_size) * 100
                                            summary_data['Отток из сети %'][cohort] = f"{percent:.1f}%"
                                        else:
                                            summary_data['Отток из сети %'][cohort] = "0.0%"
                        
                        # Создаем DataFrame
                        summary_df = pd.DataFrame(summary_data, index=sorted_periods).T
                        
                        # Отображаем таблицу
                        st.dataframe(
                            summary_df,
                            use_container_width=True
                        )
                        
                        # Добавляем стили для центрирования
                        st.markdown("""
                        <style>
                        div[data-testid="stDataFrame"] table td {
                            text-align: center !important;
                        }
                        div[data-testid="stDataFrame"] table th {
                            text-align: center !important;
                        }
                        </style>
                        """, unsafe_allow_html=True)
                    else:
                        st.info("Загрузите данные для отображения сводной таблицы")
                    
            except Exception as e:
                st.error(f"❌ Ошибка при построении матрицы: {str(e)}")
                st.exception(e)
        else:
            st.warning("⚠️ Необходимо указать столбцы для построения матрицы")
            
    except Exception as e:
        st.error(f"❌ Ошибка при загрузке файла: {str(e)}")
        st.session_state.uploaded_data = None
        st.session_state.df = None

